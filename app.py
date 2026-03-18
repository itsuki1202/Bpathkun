import streamlit as st
from PIL import Image as _PIL_Image
# ページ設定（faviconとタイトル）
_favicon = _PIL_Image.open("favicon.png")
st.set_page_config(page_title="ブライトパスくん", page_icon=_favicon, layout="wide")

import pandas as pd
import numpy as np
import io
import os
from utils.data_manager import (
    load_user_master, save_user_master, load_scoring_config, save_scoring_config,
    DATA_DIR, load_performance_data
)
from utils.calculation import calculate_scores, calculate_denominators, aggregate_team_scores, aggregate_shop_scores
from data_loader import load_data
import plotly.graph_objects as go
from master_data import PERFORMANCE_ITEMS, NEW_FORMAT_COLUMNS
from utils.auth_manager import (
    load_auth_config, save_auth_config, hash_password,
    add_user, delete_user, reset_password, log_login, load_login_log, get_all_users,
    generate_user_template_excel, bulk_add_users
)
import streamlit_authenticator as stauth
# CX分配くんる用モジュール（既存ロジックには一切たっていない）
from utils.cx_distributor import (
    load_cx_config, save_cx_config, get_default_cx_members,
    calculate_cx_distribution, SKILL_TYPES, ROLE_LABELS, ROLE_LABELS_REV,
    MANAGER_DEFAULT_PTS, MANAGER_DEFAULT_RATIOS,
    LEADER_DEFAULT_PTS, LEADER_DEFAULT_RATIOS,
)

def normalize_text(text):
    """姓名や店舗名・チーム名の空白・記号・表記揺れを徹底排除して正規化する"""
    if pd.isna(text): return ""
    import unicodedata
    import re
    t = str(text)
    # NFKC正規化（全角英数→半角、全角記号→半角、など）
    # 例：「！」→「!」「ａ」→「a」「１」→「1"
    t = unicodedata.normalize('NFKC', t)
    # 全ての空白（改行、タブ、全角スペース含む）を除去
    t = re.sub(r'[\s　]+', '', t)
    return t

def apply_robust_sorting(df, user_master, name_col, shop_col=None, team_col=None):
    """名簿順(空白・記号無視)でデータをソートし、店舗・チーム情報を名簿側で上書きする"""
    if not user_master.get('users'):
        return df
    
    df = df.copy()
    users = user_master['users']
    
    # 照合用キー（正規化済み）と名簿インデックスを生成
    user_order = {normalize_text(u['name']): i for i, u in enumerate(users)}
    # 名簿の「正規化済み店舗名」「正規化済みチーム名」を保持（元の文字列も一緒に）
    user_shops_norm = {normalize_text(u['name']): normalize_text(u.get('shop', '')) for u in users}
    user_teams_norm = {normalize_text(u['name']): normalize_text(u.get('team', '')) for u in users}
    user_shops_raw = {normalize_text(u['name']): u.get('shop', '') for u in users}
    user_teams_raw = {normalize_text(u['name']): u.get('team', '') for u in users}
    
    # スタッフ名の正規化キーを生成
    if name_col not in df.columns:
        return df
    df['_norm_name'] = df[name_col].apply(normalize_text)
    df['_master_order'] = df['_norm_name'].map(user_order).fillna(9999)
    
    # 設定画面優先：名簿に存在するスタッフの店舗・チームを名簿の値（元の文字列）で上書き
    # 名簿にないスタッフはCSV側の値をそのまま使う
    if shop_col and shop_col in df.columns:
        df[shop_col] = df.apply(
            lambda r: user_shops_raw[r['_norm_name']] if r['_norm_name'] in user_shops_raw else r[shop_col],
            axis=1
        )
    if team_col and team_col in df.columns:
        df[team_col] = df.apply(
            lambda r: user_teams_raw[r['_norm_name']] if r['_norm_name'] in user_teams_raw else r[team_col],
            axis=1
        )
    
    # 名簿順でソート（これによりgroupby().first()が正しいリーダーを返す）
    return df.sort_values('_master_order').reset_index(drop=True)

def admin_page(config, user_master, target_month=None):
    # バックナンバー対象月をヘッダーで明示
    if target_month:
        st.header(f"管理者設定 ── {target_month[:4]}年{int(target_month[4:])}月のアーカイブ")
        st.info(f"⚠️ **{target_month[:4]}年{int(target_month[4:])}月** のアーカイブを参照・編集中です。保存するとこの月のデータが上書きされます。")
    else:
        st.header("管理者設定")
    
    # タブボタンのテキストが折り返されないようにCSSを注入
    st.markdown("""
    <style>
    .stTabs [data-baseweb="tab"] {
        white-space: nowrap;
    }
    </style>
    """, unsafe_allow_html=True)

    tab1, tab2, tab3, tab4, tab_users, tab_log = st.tabs([
        "メンバー管理", "評価項目設定", "成績データ管理", "計算チェック",
        "ユーザー管理", "ログチェック"
    ])

    # ============================================================
    # タブ: ユーザー管理（管理者のみ）
    # ============================================================
    with tab_users:
        # superadmin のみアクセス可能
        if st.session_state.get("user_role") != "superadmin":
            st.warning("⛔ この画面は最高管理者のみアクセスできます。")
        else:
            st.subheader("ユーザー管理")
            _ac = load_auth_config()

            # --- 登録済みユーザー一覧 ---
            st.markdown("#### 登録済みユーザー一覧")
            _users = get_all_users(_ac)
            if _users:
                _users_df = pd.DataFrame(_users)
                _users_df.columns = ["ユーザーID", "表示名", "メール", "ロール"]
                st.dataframe(_users_df, use_container_width=True, hide_index=True)
            else:
                st.info("ユーザーが登録されていません")

            st.markdown("---")

            # --- ユーザー追加 ---
            st.markdown("#### ユーザーを追加")
            with st.form("add_user_form", clear_on_submit=True):
                _col1, _col2 = st.columns(2)
                with _col1:
                    _new_uid    = st.text_input("ユーザーID（ログイン用・英数字）", placeholder="例: tanaka01")
                    _new_name   = st.text_input("表示名", placeholder="例: 田中 太郎")
                with _col2:
                    _new_pw     = st.text_input("初期パスワード", type="password")
                    _new_role   = st.selectbox("ロール", ["viewer", "admin", "superadmin"])
                if st.form_submit_button("追加する"):
                    _ok, _msg = add_user(_ac, _new_uid, _new_name, "", _new_pw, _new_role)
                    if _ok:
                        st.success(_msg)
                        st.rerun()
                    else:
                        st.error(_msg)

            st.markdown("---")

            # --- ユーザー一括登録 ---
            st.markdown("#### 📥 ユーザー一括登録")

            # フォーマットDLボタン
            _tmpl_bytes = generate_user_template_excel()
            st.download_button(
                label="📄 登録用フォーマット(Excel)をダウンロード",
                data=_tmpl_bytes,
                file_name="user_register_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

            # アップローダー
            _bulk_file = st.file_uploader(
                "記入済みExcelをアップロード",
                type=["xlsx"],
                key="bulk_user_upload",
                help="フォーマットに従って記入したExcelをアップロード（2行目の注釈行は自動スキップ）"
            )

            if _bulk_file is not None:
                try:
                    _bulk_df = pd.read_excel(_bulk_file, header=0)
                    # 注釈行（2行目）を除外：ユーザーIDが「英数字のみ...」の行
                    _bulk_df = _bulk_df[~_bulk_df["ユーザーID"].astype(str).str.startswith("英数字")]
                    _bulk_df = _bulk_df.dropna(how="all")

                    # プレビュー表示
                    st.markdown("**📋 登録予定ユーザー（プレビュー）**")
                    _preview_df = _bulk_df[["ユーザーID", "表示名", "ロール"]].copy()
                    _preview_df["初期パスワード"] = "●●●●●●"  # パスワードはマスク表示
                    st.dataframe(_preview_df, use_container_width=True, hide_index=True)
                    st.caption(f"対象: {len(_bulk_df)} 行")

                    # 一括登録ボタン
                    if st.button("✅ 一括登録する", type="primary", use_container_width=True, key="bulk_register_btn"):
                        _bulk_ac = load_auth_config()
                        _ok_cnt, _skip_cnt, _msgs = bulk_add_users(_bulk_ac, _bulk_df)
                        if _ok_cnt > 0:
                            st.success(f"✅ {_ok_cnt} 件のユーザーを登録しました（スキップ: {_skip_cnt} 件）")
                        else:
                            st.warning(f"登録件数: 0 件（スキップ: {_skip_cnt} 件）")
                        if _msgs:
                            with st.expander("登録詳細"):
                                st.text("\n".join(_msgs))
                        if _ok_cnt > 0:
                            st.rerun()
                except Exception as _e:
                    st.error(f"Excelの読み込みに失敗しました: {_e}")
                    st.info("フォーマットDLボタンからダウンロードしたテンプレートを使用してください")

            st.markdown("---")

            # --- パスワードリセット ---
            st.markdown("#### パスワードをリセット")

            with st.form("reset_pw_form", clear_on_submit=True):
                _uid_list = [u["username"] for u in _users]
                _reset_uid = st.selectbox("対象ユーザー", _uid_list) if _uid_list else None
                _reset_pw  = st.text_input("新しいパスワード", type="password")
                if st.form_submit_button("パスワードを変更する"):
                    if _reset_uid:
                        _ok2, _msg2 = reset_password(_ac, _reset_uid, _reset_pw)
                        if _ok2:
                            st.success(_msg2)
                        else:
                            st.error(_msg2)

            st.markdown("---")

            # --- ユーザー削除 ---
            st.markdown("#### ユーザーを削除")
            with st.form("del_user_form", clear_on_submit=True):
                # superadmin ロールのユーザーは削除不可（全員保護）
                _del_uid_list = [u["username"] for u in _users if u["role"] != "superadmin"]
                _del_uid = st.selectbox("削除対象ユーザー", _del_uid_list) if _del_uid_list else None
                if st.form_submit_button("削除する"):
                    if _del_uid:
                        _ok3, _msg3 = delete_user(_ac, _del_uid)
                        if _ok3:
                            st.success(_msg3)
                            st.rerun()
                        else:
                            st.error(_msg3)

    # ============================================================
    # タブ: ログチェック（最高管理者のみ）
    # ============================================================
    with tab_log:
        # superadmin のみアクセス可能
        if st.session_state.get("user_role") != "superadmin":
            st.warning("⛔ この画面は最高管理者のみアクセスできます。")
        else:
            st.subheader("ログイン履歴")
            _log_df = load_login_log()
            if _log_df.empty:
                st.info("ログイン履歴はまだありません")
            else:
                # フィルター
                _log_col1, _log_col2 = st.columns(2)
                with _log_col1:
                    _filter_user = st.selectbox(
                        "ユーザーで絞り込み",
                        ["全員"] + list(_log_df["username"].dropna().unique())
                    )
                with _log_col2:
                    _filter_status = st.selectbox(
                        "ステータスで絞り込み",
                        ["全て", "success（成功）", "failed（失敗）"]
                    )

                _filtered = _log_df.copy()
                if _filter_user != "全員":
                    _filtered = _filtered[_filtered["username"] == _filter_user]
                if _filter_status == "success（成功）":
                    _filtered = _filtered[_filtered["status"] == "success"]
                elif _filter_status == "failed（失敗）":
                    _filtered = _filtered[_filtered["status"] == "failed"]

                # 表示列名を日本語化
                _filtered_display = _filtered.rename(columns={
                    "timestamp":    "日時",
                    "username":     "ユーザーID",
                    "display_name": "表示名",
                    "status":       "ステータス",
                    "message":      "メッセージ",
                })
                st.dataframe(_filtered_display, use_container_width=True, hide_index=True)
                st.caption(f"合計 {len(_filtered)} 件")


    
    # --- Tab 1: Member Management ---
    with tab1:
        st.subheader("メンバー管理 (User Master)")
        st.markdown("ここで設定した「店舗」「チーム」情報が、成績集計時の正となります。")
        
        # --- File Upload/Download Area ---
        col_dl, col_ul = st.columns(2)
        
        # Template Download
        with col_dl:
            # Create template dataframe
            template_df = pd.DataFrame(columns=["店舗", "チーム", "スタッフ名", "基準時間"])
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                template_df.to_excel(writer, index=False, sheet_name='UserMaster')
            
            st.download_button(
                label="登録用フォーマット(Excel)をダウンロード",
                data=buffer.getvalue(),
                file_name="member_master_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        # File Upload
        with col_ul:
            uploaded_file = st.file_uploader("メンバー情報を一括登録 (Excel)", type=["xlsx"])
            if uploaded_file is not None:
                try:
                    new_users_df = pd.read_excel(uploaded_file)
                    # Simple validation
                    required_cols = {"店舗", "チーム", "氏名"}
                    if required_cols.issubset(new_users_df.columns):
                        # Convert to JSON format and normalize keys and values
                        # Standardize columns to: name, shop, team, standard_hours
                        rename_map = {"氏名": "name", "スタッフ名": "name", "店舗": "shop", "チーム": "team", "基準時間": "standard_hours"}
                        users_df_norm = new_users_df.rename(columns=rename_map)
                        
                        # Apply row-level cleaning for store/team/name
                        def clean_row(row):
                            if 'name' in row: row['name'] = normalize_text(row['name'])
                            if 'shop' in row: row['shop'] = normalize_text(row['shop'])
                            if 'team' in row: row['team'] = normalize_text(row['team'])
                            return row
                        users_df_norm = users_df_norm.apply(clean_row, axis=1)

                        # Keep only necessary columns
                        valid_cols = ["name", "shop", "team", "standard_hours"]
                        users_df_norm = users_df_norm[[c for c in valid_cols if c in users_df_norm.columns]]
                        
                        users_list = users_df_norm.to_dict(orient='records')
                        # Keep existing shop_managers
                        current_master = load_user_master()
                        current_master["users"] = users_list
                        
                        from datetime import datetime
                        now_str = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
                        current_master["last_updated"] = now_str
                        
                        save_user_master(current_master)
                        st.success(f"メンバー情報を正規化して一括更新しました！({now_str})")
                        st.rerun()
                    else:
                        st.error(f"Excelの列名が正しくありません。{required_cols} を含めてください。")
                except Exception as e:
                    st.error(f"読み込みエラー: {e}")

        st.markdown("---")
        
        # --- Shop Manager Settings ---
        st.subheader("店舗別設定")
        st.write("店舗ごとの基本情報を設定します。")
        
        users_list = user_master.get('users', [])
        all_shops = sorted(list(set([u.get('shop') for u in users_list if u.get('shop')])))
        
        shop_managers_dict = user_master.get('shop_managers', {})
        shop_mgr_df = pd.DataFrame([
            {"店舗名": s, "店長名": shop_managers_dict.get(s, "")} 
            for s in all_shops
        ])
        
        edited_mgrs = st.data_editor(
            shop_mgr_df,
            column_config={
                "店舗名": st.column_config.TextColumn("店舗名", disabled=True),
                "店長名": st.column_config.TextColumn("店長名", help="店舗ランキング等に表示される店長名を入力してください。")
            },
            hide_index=True,
            use_container_width=True,
            key="mgr_editor"
        )

        st.markdown("---")
        # Inline Editing for Members
        st.subheader("メンバー一覧")
        st.write("▼ メンバーの追加・編集（並び順の1番目がチームリーダーとして表示されます）")
        df_users = pd.DataFrame(users_list)
        if df_users.empty:
            df_users = pd.DataFrame(columns=["店舗", "チーム", "氏名", "基準時間"])
        
        column_config = {
            "shop": st.column_config.TextColumn("店舗"),
            "team": st.column_config.TextColumn("チーム"),
            "name": st.column_config.TextColumn("スタッフ名"),
            "standard_hours": st.column_config.NumberColumn(
                "基準時間",
                help="1日あたりの標準労働時間 (デフォルト 7.5)",
                min_value=0.0,
                max_value=24.0,
                step=0.5,
                format="%.1f"
            )
        }

        # Handle column names and remove any existing role/役職 columns
        rename_map = {"店舗": "shop", "チーム": "team", "氏名": "name", "基準時間": "standard_hours"}
        df_users.rename(columns={k: v for k, v in rename_map.items() if k in df_users.columns}, inplace=True)
        if 'role' in df_users.columns: df_users.drop(columns=['role'], inplace=True)
        if '役職' in df_users.columns: df_users.drop(columns=['役職'], inplace=True)

        edited_users = st.data_editor(
            df_users, 
            column_config=column_config,
            column_order=["shop", "team", "name", "standard_hours"],
            num_rows="dynamic", 
            use_container_width=True,
            key="user_editor"
        )
        
        col_btn, col_time = st.columns([1, 2])
        with col_btn:
            save_clicked = st.button("設定をすべて保存", use_container_width=True)
        with col_time:
            last_upd = user_master.get('last_updated', 'なし')
            st.markdown(f"<div style='padding-top:5px; color:#666; font-size:14px;'>最終更新: {last_upd}</div>", unsafe_allow_html=True)

        if save_clicked:
            # Update shop managers
            new_mgrs = {normalize_text(row["店舗名"]): row["店長名"] for _, row in edited_mgrs.iterrows()}
            
            # Update users list
            clean_users = edited_users.dropna(how='all').to_dict(orient='records')
            for u in clean_users:
                u['name'] = normalize_text(u.get('name', ''))
                u['shop'] = normalize_text(u.get('shop', ''))
                u['team'] = normalize_text(u.get('team', ''))
                try:
                    u['standard_hours'] = float(u.get('standard_hours', 7.5))
                except:
                    u['standard_hours'] = 7.5
                if 'role' in u: del u['role']
            
            from datetime import datetime
            now_str = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
            
            save_user_master({
                "users": clean_users,
                "shop_managers": new_mgrs,
                "last_updated": now_str
            }, month=target_month)
            dest_label = f"{target_month[:4]}年{int(target_month[4:])}月アーカイブ" if target_month else "最新"
            st.success(f"メンバー情報を保存しました！({now_str})【保存先: {dest_label}】")
            st.rerun()

    # --- Tab 2: Scoring Config ---
    with tab2:
        st.subheader("評価項目・配点設定")
        
        # Get distinct shops for headers
        users_list_master = user_master.get('users', [])
        distinct_shops = sorted(list(set([u.get('shop') for u in users_list_master if u.get('shop')])))
        
        # --- File Upload/Download Area ---
        col_dl_score, col_ul_score = st.columns(2)
        
        # Template Data for Scoring (Base columns)
        base_template_data = {
            "項目名": ["HS新規", "スマホ総販"],
            "分母種別": ["SAITO式", "SAITO式"],
            "配点1": [60, 100], "配点2": [0, 0], "配点3": [0, 0],
            "評価軸": ["相対評価", "相対評価"],
            "獲得率目標1": [0.5, 0], "獲得率目標2": [0, 0], "獲得率目標3": [0, 0],
            "達成率目標1": [1.0, 0], "達成率目標2": [0, 0], "達成率目標3": [0, 0],
            "個人救済": [False, False], "チーム救済": [False, False]
        }
        # Add dynamic shop columns to template
        for shop in distinct_shops:
            base_template_data[f"目標_{shop}"] = [0, 0]
            
        with col_dl_score:
            template_df_score = pd.DataFrame(base_template_data)
            buffer_score = io.BytesIO()
            with pd.ExcelWriter(buffer_score, engine='openpyxl') as writer:
                template_df_score.to_excel(writer, index=False, sheet_name='ScoringConfig')
            
            st.download_button(
                label="設定用フォーマット(Excel)をダウンロード",
                data=buffer_score.getvalue(),
                file_name="scoring_config_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        with col_ul_score:
            uploaded_scoring = st.file_uploader("評価設定を一括登録 (Excel)", type=["xlsx"], key="scoring_uploader")

        # Load items from config
        items = config.get('items', [])
        
        # Mapping for Evaluation Types (Display <-> Internal)
        eval_type_map = {
            "relative": "相対評価", "absolute": "絶対評価", 
            "relative_absolute": "相対絶対", "shop_achievement": "店舗達成率",
            "individual_achievement": "個人達成",
            "team_absolute": "チーム絶対", "team_relative": "チーム相対"
        }
        eval_type_map_rev = {v: k for k, v in eval_type_map.items()}

        flat_items = []
        for item in items:
            etype_display = eval_type_map.get(item.get('evaluation_type'), item.get('evaluation_type'))
            
            flat_item = {
                "項目名": item.get('name', ''),
                "分母種別": item.get('denominator_type', 'SAITO式'),
                "配点1": item.get('weight', 0),
                "配点2": item.get('weights', {}).get('w2', 0),
                "配点3": item.get('weights', {}).get('w3', 0),
                "評価軸": etype_display,
                "獲得率目標1": item.get('targets', {}).get('acquisition_rate_1', 0.0),
                "獲得率目標2": item.get('targets', {}).get('acquisition_rate_2', 0.0),
                "獲得率目標3": item.get('targets', {}).get('acquisition_rate_3', 0.0),
                "達成率目標1": item.get('targets', {}).get('achievement_rate_1', 0.0),
                "達成率目標2": item.get('targets', {}).get('achievement_rate_2', 0.0),
                "達成率目標3": item.get('targets', {}).get('achievement_rate_3', 0.0),
                "個人救済": item.get('flags', {}).get('individual_salvage', False),
                "チーム救済": item.get('flags', {}).get('team_salvage', False),
            }
            
            # Add shop targets
            shop_targets = item.get('shop_targets', {})
            for shop in distinct_shops:
                flat_item[f"目標_{shop}"] = shop_targets.get(shop, 0)
                
            flat_items.append(flat_item)
            
        df_items = pd.DataFrame(flat_items)
        if df_items.empty:
             df_items = pd.DataFrame(columns=base_template_data.keys())
        
        # Ensure all dynamic shop columns exist in dataframe for display
        for shop in distinct_shops:
            col_name = f"目標_{shop}"
            if col_name not in df_items.columns:
                df_items[col_name] = 0

        # Process Upload
        if uploaded_scoring is not None:
            try:
                uploaded_df_sc = pd.read_excel(uploaded_scoring)
                # Basic validation: check if '項目名' exists
                if "項目名" in uploaded_df_sc.columns:
                    # Fill missing columns
                    for col in base_template_data.keys():
                        if col not in uploaded_df_sc.columns:
                            if "配点" in col or "目標" in col:
                                uploaded_df_sc[col] = 0
                            elif "救済" in col:
                                uploaded_df_sc[col] = False
                            else:
                                uploaded_df_sc[col] = ""
                    
                    # Ensure shop columns exist
                    for shop in distinct_shops:
                        col_name = f"目標_{shop}"
                        if col_name not in uploaded_df_sc.columns:
                            uploaded_df_sc[col_name] = 0
                            
                    df_items = uploaded_df_sc
                    st.success("設定ファイルを読み込みました。確認して保存してください。")
                else:
                    st.error("「項目名」列が見つかりません。")
            except Exception as e:
                st.error(f"読み込みエラー: {e}")

        # Column Config
        column_config_score = {
            "項目名": "項目名",
            "分母種別": st.column_config.SelectboxColumn(
                "分母種別", options=["SAITO式", "独自（項目名＋分母）"], required=True
            ),
            "評価軸": st.column_config.SelectboxColumn(
                "評価軸",
                options=list(eval_type_map.values()),
                required=True
            ),
            "配点1": st.column_config.NumberColumn("配点1", min_value=0, max_value=1000),
            "配点2": st.column_config.NumberColumn("配点2", min_value=0, max_value=1000),
            "配点3": st.column_config.NumberColumn("配点3", min_value=0, max_value=1000),
            "獲得率目標1": st.column_config.NumberColumn("獲得率目標1", format="%.2f"),
            "獲得率目標2": st.column_config.NumberColumn("獲得率目標2", format="%.2f"),
            "獲得率目標3": st.column_config.NumberColumn("獲得率目標3", format="%.2f"),
            "達成率目標1": st.column_config.NumberColumn("達成率目標1", format="%.2f"),
            "達成率目標2": st.column_config.NumberColumn("達成率目標2", format="%.2f"),
            "達成率目標3": st.column_config.NumberColumn("達成率目標3", format="%.2f"),
            "個人救済": "個人救済(ON/OFF)",
            "チーム救済": "チーム救済(ON/OFF)"
        }
        
        for shop in distinct_shops:
            column_config_score[f"目標_{shop}"] = st.column_config.NumberColumn(f"目標_{shop}")
        
        edited_items = st.data_editor(
            df_items, 
            column_config=column_config_score,
            num_rows="dynamic", 
            use_container_width=True
        )
        
        col_btn_sc, col_time_sc = st.columns([1, 2])
        with col_btn_sc:
            save_sc_clicked = st.button("評価設定を保存", use_container_width=True)
        with col_time_sc:
            last_upd = user_master.get('last_updated', 'なし')
            st.markdown(f"<div style='padding-top:5px; color:#666; font-size:14px;'>最終更新: {last_upd}</div>", unsafe_allow_html=True)

        if save_sc_clicked:
            new_items = []
            for _, row in edited_items.iterrows():
                internal_etype = eval_type_map_rev.get(row['評価軸'], 'relative')
                shop_targets_data = {}
                for shop in distinct_shops:
                    col_name = f"目標_{shop}"
                    if col_name in row:
                        shop_targets_data[shop] = row[col_name]

                new_item = {
                    "id": f"item_{row['項目名']}", 
                    "name": row['項目名'],
                    "weight": row['配点1'],
                    "weights": {
                        "w1": row['配点1'],
                        "w2": row['配点2'],
                        "w3": row['配点3']
                    },
                    "evaluation_type": internal_etype,
                    "denominator_type": row['分母種別'],
                    "targets": {
                        "acquisition_rate_1": row.get('獲得率目標1', 0),
                        "acquisition_rate_2": row.get('獲得率目標2', 0),
                        "acquisition_rate_3": row.get('獲得率目標3', 0),
                        "achievement_rate_1": row.get('達成率目標1', 0),
                        "achievement_rate_2": row.get('達成率目標2', 0),
                        "achievement_rate_3": row.get('達成率目標3', 0)
                    },
                    "flags": {
                        "individual_salvage": row['個人救済'],
                        "team_salvage": row['チーム救済']
                    },
                    "shop_targets": shop_targets_data
                }
                new_items.append(new_item)
            
            from datetime import datetime
            now_str = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
            user_master['last_updated'] = now_str
            save_user_master(user_master, month=target_month)
            
            save_scoring_config({"items": new_items}, month=target_month)
            dest_label = f"{target_month[:4]}年{int(target_month[4:])}月アーカイブ" if target_month else "最新"
            st.success(f"評価設定を保存しました！({now_str})【保存先: {dest_label}】")
            st.rerun()

    # --- Tab 3: Performance Data Management ---
    with tab3:
        st.subheader("成績データ一括登録")
        st.markdown("ここでアップロードしたエクセルデータが、「総合」「件数」「順位」すべてに反映されます。")
        
        col_dl_perf, col_ul_perf = st.columns(2)
        
        with col_dl_perf:
            st.markdown("**① フォーマットの取得**")
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            perf_template_df = pd.DataFrame(columns=NEW_FORMAT_COLUMNS)
            buffer_perf = io.BytesIO()
            with pd.ExcelWriter(buffer_perf, engine='openpyxl') as writer:
                perf_template_df.to_excel(writer, index=False, sheet_name='PerformanceData')
                
            st.download_button(
                label="成績データ登録用フォーマット(Excel)をダウンロード",
                data=buffer_perf.getvalue(),
                file_name="performance_data_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        with col_ul_perf:
            st.markdown("**② データのアップロード**")
            uploaded_perf = st.file_uploader("成績データをアップロード", label_visibility="collapsed", type=["csv", "xlsx"], key="perf_uploader")
            
            if uploaded_perf is not None:
                # 評価設定から現在の項目名を動的に取得し、マッピングに一時追加
                from master_data import COLUMN_MAPPING
                temp_items = config.get('items', [])
                for item in temp_items:
                    name = item.get('name')
                    if name:
                        COLUMN_MAPPING[name] = name
                
                loaded_df, err = load_data(uploaded_perf)
                if err:
                    st.error(err)
                else:
                    defined_item_names = [item.get('name') for item in temp_items if item.get('name')]
                    found_items = [col for col in defined_item_names if col in loaded_df.columns]
                    
                    if not found_items and defined_item_names:
                        st.warning(f"アップロードされたデータに現在の評価項目が見つかりません。")
                    
                    os.makedirs('data', exist_ok=True)
                    loaded_df.to_csv('data/current_performance.csv', index=False, encoding='utf-8-sig')
                    st.success("成績データをシステムに反映しました！")
                    st.session_state['use_sample'] = False
                    
        # Debug/Preview area
        if os.path.exists("data/current_performance.csv"):
             st.markdown("---")
             st.write("▼ 現在登録されているデータ (編集可能)")
             current_perf_df = pd.read_csv("data/current_performance.csv")
             edited_perf_df = st.data_editor(current_perf_df, use_container_width=True, num_rows="dynamic")
             
             if st.button("成績データの編集内容を保存"):
                 edited_perf_df.to_csv("data/current_performance.csv", index=False, encoding='utf-8-sig')
                 st.success("編集内容を保存しました！")
             
             st.markdown("---")
             st.subheader("成績分母の計算根拠 (デバッグ用)")
             if st.checkbox("計算プロセスと内訳の詳細を表示する"):
                 try:
                     df_cur = pd.read_csv("data/current_performance.csv")
                     df_denom_debug = get_denominator_df(df_cur)
                     if not df_denom_debug.empty:
                         st.write("各メンバーの「成績分母」一覧です。")
                         st.dataframe(df_denom_debug)
                     else:
                         st.info("計算可能なデータがありません。")
                 except Exception as e:
                     st.error(f"デバッグ情報の生成エラー: {e}")

        st.markdown("---")
        st.subheader("日別実績データ登録 (分母計算用)")
        st.markdown("分母計算で使用する、日別のオーダー数・端末販売数のエクセルをアップロードします。")
        
        col_dl_daily, col_ul_daily = st.columns(2)
        
        with col_dl_daily:
            st.markdown("**① 日別フォーマットの取得**")
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            daily_cols = ["販売店名", "受付担当者"] + [str(i) for i in range(1, 32)]
            daily_df = pd.DataFrame(columns=daily_cols)
            buf_daily = io.BytesIO()
            with pd.ExcelWriter(buf_daily, engine='openpyxl') as w:
                header1 = pd.DataFrame([[""]*len(daily_cols)])
                header2 = pd.DataFrame([[""]*len(daily_cols)])
                header3 = pd.DataFrame([["", ""] + [str(i) for i in range(1, 32)]])
                header4 = pd.DataFrame([daily_cols])
                final_df = pd.concat([header1, header2, header3, header4, daily_df], ignore_index=True)
                final_df.to_excel(w, index=False, header=False, sheet_name='DailyDataFormat')
            st.download_button(
                label="日別実績登録用フォーマット(Excel)をDL",
                data=buf_daily.getvalue(),
                file_name="daily_performance_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_daily_tmpl"
            )

        st.markdown("---")
        if target_month:
            # バックナンバー編集モード: 対象月を直接保存
            st.subheader(f"{target_month[:4]}年{int(target_month[4:])}月アーカイブへの保存")
            st.markdown(f"このページで行ったアップロード・編集内容を **{target_month[:4]}年{int(target_month[4:])}月** のアーカイブへ保存します。")
            if st.button(f"✅ {target_month[:4]}年{int(target_month[4:])}月アーカイブを上書き保存"):
                from utils.data_manager import archive_month_data, get_file_path
                # アーカイブの実績データを読み込み
                perf_path = get_file_path('performance_data.csv', month=target_month)
                df_to_save = pd.read_csv(perf_path) if os.path.exists(perf_path) else None
                if archive_month_data(target_month, df_to_save):
                    st.success(f"✅ {target_month[:4]}年{int(target_month[4:])}月のアーカイブを上書き保存しました！")
                else:
                    st.error("保存失敗")
        else:
            # 最新モード: 従来通りの確定・保存ボタン
            st.subheader("評価データの確定と保存 (バックナンバー作成)")
            st.markdown("現在のメンバー設定、配点設定、実績データを、指定した『年月』のアーカイブとして保存します。")
            col_fix_ym, col_fix_btn = st.columns([1, 1])
            with col_fix_ym:
                from datetime import datetime
                archive_ym = st.text_input("保存する年月 (例: 202504)", value=datetime.now().strftime("%Y%m"))
            with col_fix_btn:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                if st.button("現在の状態をバックナンバーとして確定・保存"):
                    if len(archive_ym) == 6 and archive_ym.isdigit():
                        df_to_save = pd.read_csv("data/current_performance.csv") if os.path.exists("data/current_performance.csv") else None
                        from utils.data_manager import archive_month_data
                        if archive_month_data(archive_ym, df_to_save):
                            st.success(f"✅ {archive_ym} のデータを保存しました！")
                        else:
                            st.error("保存失敗")
                    else:
                        st.error("YYYYMM形式で入力してください。")
            
        with col_ul_daily:
            st.markdown("**② 日別データのアップロード**")
            st.caption("■ 日別オーダー数")
            uploaded_orders = st.file_uploader("日別オーダー数", label_visibility="collapsed", type=["xlsx"], key="order_uploader")
            if uploaded_orders:
                from utils.daily_loader import load_daily_data
                df_orders, _ = load_daily_data(uploaded_orders, val_name="order_count")
                if df_orders is not None:
                    os.makedirs('data', exist_ok=True)
                    df_orders.to_csv('data/daily_orders.csv', index=False, encoding='utf-8-sig')
                    st.success("保存しました")
            
            st.caption("■ 日別端末販売数")
            uploaded_terms = st.file_uploader("日別端末販売数", label_visibility="collapsed", type=["xlsx"], key="term_uploader")
            if uploaded_terms:
                from utils.daily_loader import load_daily_data
                df_terms, _ = load_daily_data(uploaded_terms, val_name="terminal_count")
                if df_terms is not None:
                    os.makedirs('data', exist_ok=True)
                    df_terms.to_csv('data/daily_terminals.csv', index=False, encoding='utf-8-sig')
                    st.success("保存しました")


    # ========================================================
    # --- Tab 4: 計算チェック ---
    # ========================================================
    with tab4:
        st.header("計算チェック")
        st.markdown("設定・データ・計算の整合性を確認します。")

        # ────────────────────────────────────────────────────
        # セクション①: 設定整合性チェック
        # ────────────────────────────────────────────────────
        with st.expander("① 設定整合性チェック", expanded=True):
            items_chk = config.get('items', [])
            if not items_chk:
                st.info("評価項目が未設定です。Tab2「評価項目設定」から設定してください。")
            else:
                # 設定データの空行（nameやweightがNone）をスキップ
                valid_items_chk = [
                    it for it in items_chk
                    if it.get('name') and pd.notna(it.get('weight'))
                ]
                # 配点合計 --- weightがNone/nanの場合に対応する
                total_weight = sum(
                    float(it.get('weight') or 0)
                    for it in valid_items_chk
                )
                col_w1, col_w2 = st.columns([1, 2])
                with col_w1:
                    if abs(total_weight - 100) < 0.01:
                        st.success(f"✅ 配点合計: **{total_weight:.1f}点**（正常）")
                    else:
                        st.warning(f"⚠️ 配点合計: **{total_weight:.1f}点**（100点から外れています）")

                # --- 評価軸内訳 ---
                eval_type_label = {
                    "relative": "相対評価", "absolute": "絶対評価",
                    "relative_absolute": "相対絶対", "shop_achievement": "店舗達成率",
                    "individual_achievement": "個人達成",
                    "team_absolute": "チーム絶対", "team_relative": "チーム相対"
                }
                axis_summary = {}
                for it in valid_items_chk:
                    et = it.get('evaluation_type', 'relative')
                    lbl = eval_type_label.get(et, et)
                    if lbl not in axis_summary:
                        axis_summary[lbl] = {"件数": 0, "配点合計": 0.0, "項目名": []}
                    axis_summary[lbl]["件数"] += 1
                    axis_summary[lbl]["配点合計"] += float(it.get('weight') or 0)
                    axis_summary[lbl]["項目名"].append(it.get('name') or '')
                axis_rows = [
                    {"評価軸": lbl, "件数": v["件数"], "配点合計": v["配点合計"], "対象項目": "、".join([str(x) for x in v["項目名"] if x is not None and x != ''])}
                    for lbl, v in axis_summary.items()
                ]
                st.markdown("**評価軸内訳**")
                st.dataframe(pd.DataFrame(axis_rows), use_container_width=True, hide_index=True)

                # --- 目標値空欄チェック ---
                target_required_types = ['shop_achievement', 'individual_achievement']
                missing_targets = []
                for it in valid_items_chk:
                    if it.get('evaluation_type') in target_required_types:
                        shop_targets = it.get('shop_targets', {})
                        if not shop_targets or all(float(v) == 0 for v in shop_targets.values()):
                            missing_targets.append(it.get('name') or '')
                if missing_targets:
                    st.error(f"❌ 店舗目標が未設定の項目（店舗達成率/個人達成 なのに目標=0）: **{'、'.join([x for x in missing_targets if x])}**")
                else:
                    st.success("✅ 店舗目標の設定：問題なし")

                # --- SAITO式分母データのチェック ---
                uses_saito = any(it.get('denominator_type', 'SAITO式') == 'SAITO式' for it in valid_items_chk)
                if uses_saito:
                    has_orders = os.path.exists('data/daily_orders.csv')
                    has_terms = os.path.exists('data/daily_terminals.csv')
                    if not has_orders and not has_terms:
                        st.warning("⚠️ SAITO式分母を使う項目がありますが、日別実績データ（オーダー数・端末販売数）が未登録です。Tab3から登録してください。")
                    elif not has_orders:
                        st.warning("⚠️ 日別オーダー数が未登録です。")
                    elif not has_terms:
                        st.warning("⚠️ 日別端末販売数が未登録です。")
                    else:
                        st.success("✅ 日別実績データ（SAITO式分母用）：登録済み")

        # ────────────────────────────────────────────────────
        # セクション②: データ引用確認
        # ────────────────────────────────────────────────────
        with st.expander("② データ引用確認（設定項目名 vs 成績CSV列名）", expanded=False):
            perf_csv_path = 'data/current_performance.csv'
            if not os.path.exists(perf_csv_path):
                st.info("成績データが未登録です。Tab3からアップロードしてください。")
            elif not items_chk:
                st.info("評価項目が未設定です。")
            else:
                df_perf_chk = pd.read_csv(perf_csv_path)
                from utils.calculation import normalize_col_name as _ncn_chk
                norm_to_raw_chk = {_ncn_chk(c): c for c in df_perf_chk.columns}

                mapping_rows = []
                for it in items_chk:
                    iname = it.get('name', '')
                    norm_iname = _ncn_chk(iname)
                    matched_col = norm_to_raw_chk.get(norm_iname)
                    status = "✅ 一致" if matched_col else "❌ CSVに列が見つからない"
                    mapping_rows.append({
                        "設定項目名": iname,
                        "CSVの対応列名": matched_col if matched_col else "（未一致）",
                        "状態": status
                    })
                df_mapping = pd.DataFrame(mapping_rows)
                st.dataframe(df_mapping, use_container_width=True, hide_index=True)

                unmapped = df_mapping[df_mapping["状態"].str.startswith("❌")]
                if not unmapped.empty:
                    st.error(f"❌ {len(unmapped)}件の項目がCSVの列と一致しません。項目名・CSV列名を確認してください。")
                else:
                    st.success("✅ 全ての設定項目がCSVの列と一致しています。")

        # ────────────────────────────────────────────────────
        # セクション③: メンバー一致チェック
        # ────────────────────────────────────────────────────
        with st.expander("③ メンバー一致チェック（名簿 vs 成績CSV・リーダー/店長確認）", expanded=True):
            users_chk = user_master.get('users', [])
            if not users_chk:
                st.info("メンバー名簿が未設定です。Tab1から設定してください。")
            elif not os.path.exists('data/current_performance.csv'):
                st.info("成績データが未登録です。Tab3からアップロードしてください。")
            else:
                df_perf_m = pd.read_csv('data/current_performance.csv')
                name_col_m = next((c for c in ['スタッフ名', '氏名', 'Name'] if c in df_perf_m.columns), None)

                # --- 名前の正規化突合 ---
                master_norm_map = {normalize_text(u['name']): u['name'] for u in users_chk}
                master_norms = set(master_norm_map.keys())

                if name_col_m:
                    csv_names_raw = df_perf_m[name_col_m].dropna().unique().tolist()
                    csv_norm_map = {normalize_text(n): n for n in csv_names_raw}
                    csv_norms = set(csv_norm_map.keys())

                    # 名簿にいてCSVにいない
                    only_master = master_norms - csv_norms
                    # CSVにいて名簿にいない
                    only_csv = csv_norms - master_norms
                    # 一致
                    matched = master_norms & csv_norms

                    col_m1, col_m2, col_m3 = st.columns(3)
                    with col_m1:
                        if matched:
                            st.success(f"✅ 一致: {len(matched)}名")
                        else:
                            st.warning("一致するメンバーなし")
                    with col_m2:
                        if only_master:
                            st.warning(f"⚠️ 名簿のみ（CSVに未収録）: {len(only_master)}名")
                        else:
                            st.success("✅ 名簿 → CSV: 全員収録")
                    with col_m3:
                        if only_csv:
                            st.warning(f"⚠️ CSVのみ（名簿に未登録）: {len(only_csv)}名")
                        else:
                            st.success("✅ CSV → 名簿: 全員登録済み")

                    # 詳細テーブル
                    detail_rows = []
                    for nm, raw in master_norm_map.items():
                        csv_raw = csv_norm_map.get(nm)
                        status_m = "✅ 一致" if csv_raw else "⚠️ CSVに未収録"
                        detail_rows.append({"名簿の名前": raw, "正規化後": nm, "CSVの名前": csv_raw or "―", "状態": status_m})
                    for nm, raw in csv_norm_map.items():
                        if nm not in master_norm_map:
                            detail_rows.append({"名簿の名前": "―", "正規化後": nm, "CSVの名前": raw, "状態": "❌ 名簿に未登録"})
                    st.markdown("**メンバー突合詳細**")
                    df_detail = pd.DataFrame(detail_rows).sort_values("状態")
                    st.dataframe(df_detail, use_container_width=True, hide_index=True)

                st.markdown("---")

                # --- チームリーダー確認 ---
                st.markdown("**チームリーダー確認（名簿順1位が表示画面のリーダーになります）**")
                team_leader_rows = []
                seen_teams = {}
                for u in users_chk:
                    team_key = normalize_text(u.get('team', ''))
                    if team_key and team_key not in seen_teams:
                        seen_teams[team_key] = u.get('name', '')
                for team_raw, leader_name in seen_teams.items():
                    # チームに所属するメンバー数も取得
                    members = [u['name'] for u in users_chk if normalize_text(u.get('team', '')) == team_raw]
                    # 対応する店舗を確認
                    shop_name = next((u.get('shop', '') for u in users_chk if normalize_text(u.get('team', '')) == team_raw), '')
                    team_leader_rows.append({
                        "店舗": shop_name,
                        "チーム(正規化後)": team_raw,
                        "リーダー（名前）": leader_name,
                        "チーム人数": len(members),
                        "メンバー一覧": "、".join(members)
                    })
                if team_leader_rows:
                    st.dataframe(pd.DataFrame(team_leader_rows), use_container_width=True, hide_index=True)
                else:
                    st.info("チーム情報がありません。")

                st.markdown("---")

                # --- 店長確認 ---
                st.markdown("**店長確認（ランキング画面に表示される店長名）**")
                shop_managers_chk = user_master.get('shop_managers', {})
                all_shops_chk = sorted(list(set([u.get('shop', '') for u in users_chk if u.get('shop')])))
                shop_mgr_rows = []
                for s in all_shops_chk:
                    mgr = shop_managers_chk.get(normalize_text(s), shop_managers_chk.get(s, ''))
                    status_mgr = "✅ 設定済み" if mgr else "⚠️ 未設定"
                    shop_mgr_rows.append({"店舗名": s, "設定されている店長名": mgr or "（未設定）", "状態": status_mgr})
                if shop_mgr_rows:
                    st.dataframe(pd.DataFrame(shop_mgr_rows), use_container_width=True, hide_index=True)
                    unset_shops = [r["店舗名"] for r in shop_mgr_rows if r["状態"].startswith("⚠️")]
                    if unset_shops:
                        st.warning(f"⚠️ 以下の店舗の店長が未設定です: **{'、'.join(unset_shops)}**　→ Tab1「店舗別設定」で入力してください。")
                    else:
                        st.success("✅ 全店舗に店長が設定されています。")
                else:
                    st.info("店舗情報がありません。")

        # ────────────────────────────────────────────────────
        # セクション④: 計算詳細監査
        # ────────────────────────────────────────────────────
        with st.expander("④ 計算詳細監査（スコア計算の根拠を確認）", expanded=False):
            if not os.path.exists('data/current_performance.csv'):
                st.info("成績データが未登録です。Tab3からアップロードしてください。")
            elif not items_chk:
                st.info("評価項目が未設定です。")
            else:
                from utils.calculation import get_calculation_audit_df, calculate_scores, normalize_col_name as _ncn_audit
                df_audit_raw = pd.read_csv('data/current_performance.csv')
                df_denom_audit = get_denominator_df(df_audit_raw)

                # 実際に使われているスコアを calculate_scores() で取得
                # get_calculation_audit_df は絶対評価系のみ対応。
                # 相対評価・相対絶対のスコアは calculate_scores の結果で上書きする。
                scored_df_audit = calculate_scores(df_audit_raw, config, df_denom_audit)

                df_audit = get_calculation_audit_df(df_audit_raw, config, df_denom_audit)

                # calculate_scores の実スコアを監査テーブルにマージ
                if not df_audit.empty and not scored_df_audit.empty:
                    name_col_audit = next((c for c in ['スタッフ名', '氏名', 'Name'] if c in scored_df_audit.columns), None)
                    if name_col_audit:
                        # {(スタッフ名, 項目名): 実スコア} の辞書を構築
                        real_score_map = {}
                        for _item in config.get('items', []):
                            _iname = _item.get('name', '')
                            if not _iname: continue
                            _norm_col = f"{_ncn_audit(_iname)}_score"
                            if _norm_col not in scored_df_audit.columns: continue
                            for _, _srow in scored_df_audit.iterrows():
                                _sn = str(_srow[name_col_audit]).strip()
                                real_score_map[(_sn, _iname)] = float(_srow[_norm_col])

                        # df_audit の最終スコア列を実スコアで上書き
                        def _get_real_score(row):
                            key = (str(row['対象者']).strip(), str(row['項目名']).strip())
                            return real_score_map.get(key, row['最終スコア'])
                        df_audit['最終スコア'] = df_audit.apply(_get_real_score, axis=1)

                        # 実際の Total_Score も展示用に付与
                        if 'Total_Score' in scored_df_audit.columns:
                            _total_map = scored_df_audit.set_index(name_col_audit)['Total_Score'].to_dict()
                            df_audit['実合計スコア'] = df_audit['対象者'].map(
                                lambda n: _total_map.get(str(n).strip(), None)
                            )

                if df_audit.empty:
                    st.info("監査データが生成できません。データと設定を確認してください。")
                else:
                    # --- 配点合計比較 ---
                    # 配点合計（有効な項目のみ）― nanを防ぐ
                    valid_items_for_audit = [it for it in items_chk if it.get('name') and pd.notna(it.get('weight'))]
                    max_possible = sum(float(it.get('weight') or 0) for it in valid_items_for_audit)
                    if '最終スコア' in df_audit.columns and '対象者' in df_audit.columns:
                        person_totals = df_audit.groupby('対象者')['最終スコア'].sum().reset_index()
                        person_totals.columns = ['スタッフ名', '計算合計スコア']
                        person_totals['設定配点合計(最大値)'] = max_possible
                        person_totals['過超フラグ'] = person_totals['計算合計スコア'] > max_possible + 0.01
                        over_count = person_totals['過超フラグ'].sum()
                        if over_count > 0:
                            st.error(f"❌ {over_count}名のスコアが設定配点合計（{max_possible:.1f}点）を超過しています。")
                        else:
                            st.success(f"✅ 全スタッフのスコアが設定配点合計（{max_possible:.1f}点）以内に収まっています。")

                    st.markdown("---")

                    # --- フィルタパネル（店舗→チーム→スタッフの連動フィルタ） ---
                    frow1_c1, frow1_c2, frow1_c3 = st.columns(3)
                    with frow1_c1:
                        # 店舗フィルタ
                        _all_shops_audit = sorted([s for s in df_audit['店舗'].dropna().unique().tolist() if s and s != '-']) if '店舗' in df_audit.columns else []
                        sel_shop = st.selectbox("店舗で絞り込み", ["（全店舗）"] + _all_shops_audit, key="audit_shop_filter")
                    with frow1_c2:
                        # チームフィルタ（店舗連動）
                        _team_base = df_audit[df_audit['店舗'] == sel_shop] if (sel_shop != "（全店舗）" and '店舗' in df_audit.columns) else df_audit
                        _all_teams_audit = sorted([t for t in _team_base['チーム'].dropna().unique().tolist() if t and t != '-']) if 'チーム' in df_audit.columns else []
                        sel_team = st.selectbox("チームで絞り込み", ["（全チーム）"] + _all_teams_audit, key="audit_team_filter")
                    with frow1_c3:
                        # スタッフフィルタ（チーム連動）
                        _staff_base = _team_base[_team_base['チーム'] == sel_team] if (sel_team != "（全チーム）" and 'チーム' in df_audit.columns) else _team_base
                        _all_staff_audit = sorted(_staff_base['対象者'].dropna().unique().tolist()) if '対象者' in _staff_base.columns else []
                        sel_staff = st.selectbox("スタッフで絞り込み", ["（全員）"] + _all_staff_audit, key="audit_staff_filter")

                    frow2_c1, frow2_c2 = st.columns([2, 1])
                    with frow2_c1:
                        all_items_audit = ["（全項目）"] + sorted(df_audit['項目名'].unique().tolist()) if '項目名' in df_audit.columns else ["（全項目）"]
                        sel_item = st.selectbox("項目で絞り込み", all_items_audit, key="audit_item_filter")
                    with frow2_c2:
                        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                        show_zero_only = st.checkbox("0点の項目のみ表示", key="audit_zero_filter")

                    # --- フィルタ適用 ---
                    df_audit_disp = df_audit.copy()
                    if sel_shop != "（全店舗）" and '店舗' in df_audit_disp.columns:
                        df_audit_disp = df_audit_disp[df_audit_disp['店舗'] == sel_shop]
                    if sel_team != "（全チーム）" and 'チーム' in df_audit_disp.columns:
                        df_audit_disp = df_audit_disp[df_audit_disp['チーム'] == sel_team]
                    if sel_staff != "（全員）":
                        df_audit_disp = df_audit_disp[df_audit_disp['対象者'] == sel_staff]
                    if sel_item != "（全項目）":
                        df_audit_disp = df_audit_disp[df_audit_disp['項目名'] == sel_item]
                    if show_zero_only:
                        df_audit_disp = df_audit_disp[df_audit_disp['最終スコア'] == 0]

                    st.markdown(f"**計算過程テーブル**（{len(df_audit_disp)}件表示）")
                    st.dataframe(
                        df_audit_disp,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "達成率(%)": st.column_config.NumberColumn("達成率(%)", format="%.1f"),
                            "実績値": st.column_config.NumberColumn("実績値", format="%.2f"),
                            "計算分母": st.column_config.NumberColumn("計算分母", format="%.3f"),
                            "計算ターゲット": st.column_config.NumberColumn("計算ターゲット", format="%.2f"),
                            "最終スコア": st.column_config.NumberColumn("最終スコア", format="%.2f"),
                        }
                    )

                    # --- スタッフ別スコア合計サマリ（店舗・チーム列含む） ---
                    if '対象者' in df_audit.columns and '最終スコア' in df_audit.columns:
                        st.markdown("---")
                        st.markdown("**スタッフ別スコア合計サマリ**")
                        # グループキーに店舗・チームを追加
                        _grp_cols = []
                        if '店舗' in df_audit.columns:
                            _grp_cols.append('店舗')
                        if 'チーム' in df_audit.columns:
                            _grp_cols.append('チーム')
                        _grp_cols.append('対象者')
                        summary_df = df_audit.groupby(_grp_cols, dropna=False)['最終スコア'].sum().reset_index()
                        summary_df = summary_df.rename(columns={'対象者': 'スタッフ名', '最終スコア': '計算合計スコア'})
                        summary_df['設定配点合計'] = max_possible
                        summary_df['スコア率(%)'] = (summary_df['計算合計スコア'] / max_possible * 100).round(1)
                        # 店舗→チーム→スコア降順でソート
                        _sort_cols = [c for c in ['店舗', 'チーム', '計算合計スコア'] if c in summary_df.columns]
                        _sort_asc  = [True if c in ('店舗', 'チーム') else False for c in _sort_cols]
                        summary_df = summary_df.sort_values(_sort_cols, ascending=_sort_asc).reset_index(drop=True)
                        st.dataframe(
                            summary_df,
                            use_container_width=True,
                            hide_index=True,
                            column_config={
                                "計算合計スコア": st.column_config.NumberColumn("計算合計スコア", format="%.2f"),
                                "スコア率(%)": st.column_config.ProgressColumn(
                                    "スコア率(%)", min_value=0, max_value=100, format="%.1f%%"
                                )
                            }
                        )

        # ────────────────────────────────────────────────────
        # セクション⑤: 整合性チェック（個人 vs 店舗スコア）
        # ────────────────────────────────────────────────────
        with st.expander("⑤ 整合性チェック（個人スコア平均 vs 店舗集計スコア）", expanded=False):
            st.markdown("""
            **目的**: 個人スコアの店舗平均と、店舗集計スコアを独立した視点で比較します。  
            大きな乖離（差が配点の20%超）がある場合、計算ロジックの非対称バグを検出できます。
            """)
            if not os.path.exists('data/current_performance.csv'):
                st.info("成績データが未登録です。Tab3からアップロードしてください。")
            elif not items_chk:
                st.info("評価項目が未設定です。")
            else:
                df_ic = pd.read_csv('data/current_performance.csv')
                df_denom_ic = get_denominator_df(df_ic)
                # 個人スコア計算
                scored_ic = calculate_scores(df_ic, config, df_denom_ic)
                # 店舗集計スコア計算
                config_ic = dict(config)
                config_ic['_user_master'] = user_master
                shop_ic = aggregate_shop_scores(scored_ic, df_ic, config_ic)

                from utils.calculation import normalize_col_name as _ncn_ic
                shop_col_ic = '店舗名' if '店舗名' in scored_ic.columns else 'Shop'

                check_rows = []
                has_alert = False
                for it in items_chk:
                    iname = it.get('name', '')
                    if not iname: continue
                    w = float(it.get('weight', 0))
                    if w <= 0: continue
                    score_col_ic = f"{_ncn_ic(iname)}_score"
                    if score_col_ic not in scored_ic.columns: continue
                    if shop_ic is None or shop_ic.empty or score_col_ic not in shop_ic.columns: continue

                    shops_list = shop_ic['店舗名'].tolist() if '店舗名' in shop_ic.columns else []
                    for shop_name in shops_list:
                        ind_scores = scored_ic[scored_ic[shop_col_ic] == shop_name][score_col_ic].dropna()
                        if ind_scores.empty: continue
                        ind_avg = ind_scores.mean()

                        shop_row = shop_ic[shop_ic['店舗名'] == shop_name]
                        if shop_row.empty: continue
                        shop_score = float(shop_row.iloc[0][score_col_ic])

                        diff = abs(ind_avg - shop_score)
                        threshold = w * 0.20
                        status = "⚠️ 要確認" if diff > threshold else "✅ 正常"
                        if diff > threshold:
                            has_alert = True

                        check_rows.append({
                            "店舗": shop_name,
                            "項目名": iname,
                            "個人平均スコア": round(ind_avg, 2),
                            "店舗集計スコア": round(shop_score, 2),
                            "差異": round(diff, 2),
                            "配点": w,
                            "閾値(配点20%)": round(threshold, 2),
                            "判定": status,
                        })

                if check_rows:
                    df_check = pd.DataFrame(check_rows)
                    if has_alert:
                        st.error("❌ 乖離が大きい項目が検出されました（差異 > 配点の20%）。ロジックを確認してください。")
                    else:
                        st.success("✅ 全項目の個人平均スコアと店舗集計スコアは整合しています。")
                    st.dataframe(
                        df_check,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "個人平均スコア": st.column_config.NumberColumn(format="%.2f"),
                            "店舗集計スコア": st.column_config.NumberColumn(format="%.2f"),
                            "差異": st.column_config.NumberColumn(format="%.2f"),
                        }
                    )
                else:
                    st.info("比較データが生成できません。")


def get_denominator_df(df):

    """ Helper to calculate denominators on the fly processing local CSVs """
    import os
    if os.path.exists('data/daily_orders.csv'):
        df_orders = pd.read_csv('data/daily_orders.csv')
    else:
        df_orders = pd.DataFrame()
        
    if os.path.exists('data/daily_terminals.csv'):
        df_terms = pd.read_csv('data/daily_terminals.csv')
    else:
        df_terms = pd.DataFrame()
        
    user_master = load_user_master()
    df_users = pd.DataFrame(user_master.get('users', []))
    
    return calculate_denominators(df, df_orders, df_terms, df_users)


# --- Page Functions ---

def ranking_page(df, target_month=None):
    st.header("順位表 (Ranking)")
    
    if df is None:
        st.info("データをアップロードしてください。")
        return

    # Load Config（バックナンバー対応: target_monthが指定された月の設定を使用）
    config = load_scoring_config(month=target_month)
    if not config or not config.get('items'):
        st.warning("評価設定がありません。管理者メニューから設定してください。")
        st.dataframe(df)
        return

    # Calculate Denominators
    df_denom = get_denominator_df(df)
    
    # Calculate Scores with weights
    scored_df = calculate_scores(df, config, df_denom)
    
    # --- Robust Merge with User Master ---
    user_master = load_user_master(month=target_month)
    if user_master.get('users'):
        df_master = pd.DataFrame(user_master['users'])
        
        # Identify columns (Standardized names in data_loader.py)
        name_col_score = 'スタッフ名'
        shop_col_score = '店舗名'
        team_col_score = 'チーム名'
        
        # Merge if possible
        if name_col_score in scored_df.columns and 'name' in df_master.columns:
            # Perform merge (keep all score rows)
            merged_df = pd.merge(scored_df, df_master[['name', 'shop', 'team']], 
                               left_on=name_col_score, right_on='name', 
                               how='left', suffixes=('', '_master'))
            
            # Prefer Master Data (shop_master) > Uploaded Data (shop_col_score)
            # Create standardized '店舗' and 'チーム' columns
            
            # Shop Logic
            if 'shop' in merged_df.columns:
                # If master has shop, use it. If nan, fallback to upload.
                merged_df['店舗名'] = merged_df['shop'].fillna(merged_df[shop_col_score] if shop_col_score in merged_df.columns else pd.NA)
            elif shop_col_score in merged_df.columns:
                merged_df['店舗名'] = merged_df[shop_col_score]
            else:
                merged_df['店舗名'] = "未所属"
                
            # Team Logic
            if 'team' in merged_df.columns:
                merged_df['チーム名'] = merged_df['team'].fillna(merged_df[team_col_score] if team_col_score in merged_df.columns else pd.NA)
            elif team_col_score in merged_df.columns:
                merged_df['チーム名'] = merged_df[team_col_score]
            else:
                merged_df['チーム名'] = "未所属"
                
            # Clean up redundant columns if desired, but keeping them for debug is safer.
            # Update scored_df
            scored_df = merged_df
        # Apply robust sorting and master data override to the scored dataframe
        scored_df = apply_robust_sorting(scored_df, user_master, name_col_score, shop_col_score, team_col_score)
        
        # Ensure standardized columns exist after potential override
        scored_df['店舗名'] = scored_df[shop_col_score]
        scored_df['チーム名'] = scored_df[team_col_score]

    # Ensure items exist
    items = config.get('items', [])
    item_names = [item.get('name') for item in items if item.get('name')]
    
    name_col = 'スタッフ名' if 'スタッフ名' in scored_df.columns else 'Name'
    shop_col = '店舗名' if '店舗名' in scored_df.columns else 'Shop'
    team_col = 'チーム名' if 'チーム名' in scored_df.columns else 'Team'

    # --- Main Views ---
    tab_overall, tab_category = st.tabs(["総合ランキング", "部門別ランキング"])

    # 1. Overall Ranking Tab
    with tab_overall:
        from utils.calculation import normalize_col_name as _ncn
        
        # --- 個人総合ランキング ---
        st.subheader("個人総合ランキング")
        if 'Rank' in scored_df.columns:
            fixed = [(shop_col, '店舗名'), (team_col, 'チーム名'), (name_col, 'スタッフ名')]
            score_cols_ord = [
                (f"{_ncn(item.get('name',''))}_score", item.get('name',''), item.get('evaluation_type', 'relative'), float(item.get('weight', 0)))
                for item in items
                if f"{_ncn(item.get('name',''))}_score" in scored_df.columns
            ]
            
            show_df = scored_df.sort_values("Rank").copy()
            html = build_html_table(show_df, fixed, score_cols_ord, items=items)
            st.markdown(html, unsafe_allow_html=True)
        else:
            st.error("ランキング計算エラー")

        # --- チーム総合ランキング ---
        st.subheader("チーム総合ランキング")
        # 判定を名簿優先にするため、原本も名簿順で正規化ソート
        raw_name_col = next((c for c in ['スタッフ名', '氏名', 'Name'] if c in df.columns), 'スタッフ名')
        df_for_agg = apply_robust_sorting(df, user_master, raw_name_col, '店舗名', 'チーム名')

        config['_user_master'] = user_master
        team_df = aggregate_team_scores(scored_df, df_original=df_for_agg, config=config)
        
        # Leaderをuser_masterから直接上書き（計算結果に依存しない確実な方法）
        if not team_df.empty and user_master.get('users'):
            import unicodedata as _ud2, re as _re2
            def _norm_ldr2(t):
                t = _ud2.normalize('NFKC', str(t))
                return _re2.sub(r'[\s\u3000]+', '', t)
            team_ldr_map2 = {}
            for u in user_master['users']:
                tk = _norm_ldr2(u.get('team', ''))
                if tk not in team_ldr_map2:
                    team_ldr_map2[tk] = u.get('name', '')
            t_col_ldr2 = 'チーム名' if 'チーム名' in team_df.columns else ('チーム' if 'チーム' in team_df.columns else None)
            if t_col_ldr2:
                team_df['Leader'] = team_df[t_col_ldr2].apply(lambda t: team_ldr_map2.get(_norm_ldr2(t), ''))
        
        if not team_df.empty:
            t_col_in = 'チーム名' if 'チーム名' in team_df.columns else ('チーム' if 'チーム' in team_df.columns else 'Team')
            s_col_in = '店舗名' if '店舗名' in team_df.columns else ('店舗' if '店舗' in team_df.columns else 'Shop')
            leader_col = 'Leader' if 'Leader' in team_df.columns else None
            
            fixed_t = [(s_col_in, '店舗'), (t_col_in, 'チーム')]
            if leader_col:
                fixed_t.append((leader_col, 'リーダー'))
            
            team_score_cols = [
                (f"{_ncn(item.get('name',''))}_score", item.get('name',''), item.get('evaluation_type', 'relative'), float(item.get('weight', 0)))
                for item in items
                if f"{_ncn(item.get('name',''))}_score" in team_df.columns
            ]
            
            html_t = build_html_table(team_df.sort_values("Rank"), fixed_t, team_score_cols, bg_color="#FFF5F5", items=items)
            st.markdown(html_t, unsafe_allow_html=True)
        else:
            st.info("チーム情報がありません")

        st.markdown("<div style='height:30px'></div>", unsafe_allow_html=True)

        # --- 店舗総合ランキング ---
        st.subheader("店舗総合ランキング")
        # 店長情報を渡すために config に user_master を一時的に紐付ける
        config['_user_master'] = user_master
        shop_df = aggregate_shop_scores(scored_df, df_original=df, config=config)
        if not shop_df.empty:
            s_col2 = '店舗名' if '店舗名' in shop_df.columns else ('店舗' if '店舗' in shop_df.columns else 'Shop')
            mgr_col = 'Manager' if 'Manager' in shop_df.columns else None
            
            fixed_s = [(s_col2, '店舗')]
            if mgr_col:
                fixed_s.append((mgr_col, '店長'))
            
            shop_score_cols = [
                (f"{_ncn(item.get('name',''))}_score", item.get('name',''), item.get('evaluation_type', 'relative'), float(item.get('weight', 0)))
                for item in items
                if f"{_ncn(item.get('name',''))}_score" in shop_df.columns
            ]
            
            html_s = build_html_table(shop_df.sort_values("Rank"), fixed_s, shop_score_cols, bg_color="#F5FFF5", items=items)
            st.markdown(html_s, unsafe_allow_html=True)
        else:
            st.info("店舗情報がありません")

    # 2. Category Ranking Tab
    with tab_category:
        st.subheader("部門別ランキング (Top 5)")
        if not item_names:
            st.info("評価項目が設定されていません")
        else:
            cat_tabs = st.tabs(item_names)
            for i, item_name in enumerate(item_names):
                with cat_tabs[i]:
                    if item_name not in scored_df.columns:
                        st.warning(f"「{item_name}」データがありません")
                        continue
                    
                    # 項目ごとの詳細設定を取得
                    item_conf = next((it for it in items if it.get('name') == item_name), {})
                    sc_col = f"{_ncn(item_name)}_score"
                    
                    sorted_df = scored_df.sort_values(item_name, ascending=False).head(5).copy()
                    sorted_df['Rank_Tmp'] = range(1, len(sorted_df)+1)
                    
                    # 部門別なので、その項目の実績値とスコアを見せる
                    fixed_c = [(shop_col, '店舗'), (team_col, 'チーム'), (name_col, 'スタッフ名'), (item_name, '実績')]
                    # score_cols にその項目のスコアを1つだけ入れる
                    score_cols_c = [(sc_col, item_name, item_conf.get('evaluation_type','-'), float(item_conf.get('weight',0)))]
                    
                    # Rank 列を一時的に上書きして表示
                    orig_rank = sorted_df['Rank'].copy() if 'Rank' in sorted_df.columns else None
                    sorted_df['Rank'] = sorted_df['Rank_Tmp']
                    
                    html_c = build_html_table(sorted_df, fixed_c, score_cols_c)
                    st.markdown(html_c, unsafe_allow_html=True)



# =========================================================
# 共通テーブル表示関数（ヘッダー固定・3段対応）
# =========================================================
def build_html_table(df, fixed_cols, score_cols_ordered, totals_label="合計", bg_color=None, show_rank=True, items=None):
    """
    df: 表示データフレーム
    fixed_cols: [(内部カラム名, 表示名)] 固定列リスト
    score_cols_ordered: [(score_col, item_name, eval_type, weight)] 項目列リスト
    items: 全項目設定（合計配点計算用）
    """
    # Refined CSS for structural Header
    sticky_css = """<style>
.report-container { overflow: auto; overflow-x: scroll; -webkit-overflow-scrolling: touch; height: 600px; border: 1px solid #ddd; border-radius: 8px; position: relative; }
.report-table { border-collapse: separate; border-spacing: 0; width: 100%; background-color: white; }
/* ヘッダー固定: padding不要・下揃えベースに変更 */
.report-table thead th { position: sticky; top: 0; background: #f8f9fa; z-index: 30; border-bottom: 2px solid #dee2e6; border-right: 1px solid #eee; padding: 0; color: #202124; vertical-align: bottom; }
/* 横スクロール固定列 */
.sticky-col { position: sticky; background: white !important; z-index: 5; border-right: 1px solid #ddd !important; }
/* 左上コーナー（固定列ヘッダー） */
th.sticky-col { z-index: 40 !important; background: #f8f9fa !important; }
.report-table td { padding: 8px 12px; border-bottom: 1px solid #eee; border-right: 1px solid #eee; font-size: 13px; white-space: nowrap; }
.report-table tbody tr:hover { background-color: #f1f3f4; }
.report-table tbody tr:hover td.sticky-col { background-color: #f1f3f4 !important; }
.text-right { text-align: right; }
.text-left { text-align: left; }
.font-bold { font-weight: bold; }
/* ヘッダー3段構造: flexboxで「評価軸」「点数」を常に底辺に揃える */
.header-wrapper {
  display: flex;
  flex-direction: column;
  justify-content: flex-end;
  align-items: center;
  min-height: 72px;
  min-width: 90px;
  padding: 6px 10px 6px 10px;
  text-align: center;
  box-sizing: border-box;
}
/* 項目名: 上部スペースを占有し、テキストは下揃え・自然な折り返しを許可 */
.header-item {
  font-weight: bold;
  font-size: 11px;
  line-height: 1.4;
  white-space: normal;
  word-break: normal;
  overflow-wrap: anywhere;
  flex-grow: 1;
  display: flex;
  align-items: flex-end;
  justify-content: center;
  width: 100%;
  margin-bottom: 2px;
}
/* 評価軸: 1行固定・縦位置統一 */
.header-type {
  font-size: 10px;
  color: #666666;
  white-space: nowrap !important;
  flex-shrink: 0;
  min-width: max-content;
  margin-top: 1px;
}
/* 点数: 1行固定・青色・太字 */
.header-weight {
  font-size: 12px;
  color: #0066cc;
  font-weight: bold;
  white-space: nowrap !important;
  flex-shrink: 0;
  min-width: max-content;
  margin-top: 1px;
}
/* PC: shop-shortは非表示・shop-fullは表示 */
.shop-short { display: none; }
.shop-full { display: inline; }
/* ===== モバイル専用スタイル（768px以下のみ適用・PCには影響なし） ===== */
@media screen and (max-width: 768px) {
  /* テーブルのフォントサイズ・余白を縮小 */
  .report-table td { font-size: 10px !important; padding: 4px 5px !important; }
  /* 固定列（店舗・チーム・スタッフ名）の幅を縮小して数値列を見やすく */
  .sticky-col { min-width: 60px !important; max-width: 80px !important; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
  th.sticky-col { min-width: 60px !important; }
  /* 固定列のleft位置をモバイル幅（各60px）に合わせて上書き（位置ズレ修正） */
  .report-table td.sticky-col:nth-child(1), .report-table th.sticky-col:nth-child(1) { left: 0px !important; }
  .report-table td.sticky-col:nth-child(2), .report-table th.sticky-col:nth-child(2) { left: 60px !important; }
  .report-table td.sticky-col:nth-child(3), .report-table th.sticky-col:nth-child(3) { left: 120px !important; }
  /* スコア列の幅も縮小 */
  .report-table th { min-width: 60px !important; width: auto !important; }
  /* ヘッダーのテキストを小さく */
  .header-wrapper { min-width: 60px !important; padding: 3px 4px !important; }
  .header-item { font-size: 8px !important; }
  .header-type { font-size: 7px !important; }
  .header-weight { font-size: 8px !important; }
  /* コンテナ高さをモバイル向けに最適化 */
  .report-container { height: 420px !important; }
  /* モバイルでは店舗名を短縮表示（「ドコモショップ」を除去） */
  .shop-full { display: none; }
  .shop-short { display: inline; }
}
</style>"""
    
    etype_jp = {
        'relative': '相対評価', 'absolute': '絶対評価',
        'relative_absolute': '相対絶対', 
        'team_relative': 'チーム相対', 'team_absolute': 'チーム絶対', 
        'shop_achievement': '店舗達成率', 'individual_achievement': '個人達成'
    }

    # 合計配点 = 実際に表示している列のweightの合計（itemsの全項目ではなく表示列のみ）
    total_weight = sum(float(w or 0) for _, _, _, w in score_cols_ordered)

    html = [sticky_css, '<div class="report-container"><table class="report-table">']

    # 固定列の位置計算
    col_offsets = []
    current_offset = 0
    for i, (col, _) in enumerate(fixed_cols):
        col_offsets.append(current_offset)
        if "スタッフ" in col or "Name" in col or "名" in col or "店舗" in col or "Shop" in col:
            current_offset += 160
        else:
            current_offset += 100

    # ------ ヘッダー行 (1段集約・構造化) ------
    html.append('<thead>')
    html.append('<tr>')
    
    # 固定列 (店舗、チーム、スタッフ名など)
    for i, (col, disp) in enumerate(fixed_cols):
        min_w = 160 if ("スタッフ" in col or "Name" in col or "名" in col or "店舗" in col or "Shop" in col) else 100
        style = f"left: {col_offsets[i]}px; min-width: {min_w}px;"
        # header-wrapperで高さを他の列と統一（テキストは左揃え）
        html.append(
            f'<th class="sticky-col" style="{style}">'
            f'<div class="header-wrapper" style="align-items:flex-start;">'
            f'<div class="header-item" style="font-size:14px; justify-content:flex-start;">{disp}</div>'
            f'</div></th>'
        )
    
    # 項目列 (構造化された中央揃え)
    for _, iname, etype, w in score_cols_ordered:
        jp = etype_jp.get(etype, etype)
        w_val = int(w) if w == int(w) else w
        header_html = f'<div class="header-wrapper">' \
                      f'<div class="header-item">{iname}</div>' \
                      f'<div class="header-type">{jp}</div>' \
                      f'<div class="header-weight">{w_val}点</div>' \
                      f'</div>'
        html.append(f'<th style="min-width: 110px; width: 110px;">{header_html}</th>')
    
    # 合計・順位
    total_header = f'<div class="header-wrapper">' \
                   f'<div class="header-item">{totals_label}</div>' \
                   f'<div style="height:10px;"></div>' \
                   f'<div class="header-weight">{int(total_weight) if total_weight == total_weight else 0}点</div>' \
                   f'</div>'
    html.append(f'<th style="min-width: 90px;">{total_header}</th>')
    
    if show_rank:
        rank_header = f'<div class="header-wrapper">' \
                      f'<div class="header-item">順位</div>' \
                      f'</div>'
        html.append(f'<th style="min-width: 70px;">{rank_header}</th>')
        
    html.append('</tr>')
    html.append('</thead>')

    # ------ データ行 ------
    html.append('<tbody>')
    row_parent_bg = f"background:{bg_color};" if bg_color else ""
    for _, row in df.iterrows():
        html.append(f'<tr style="{row_parent_bg}">')
        for i, (col, _) in enumerate(fixed_cols):
            v = row.get(col, '')
            min_w = 160 if ("スタッフ" in col or "Name" in col or "名" in col or "店舗" in col or "Shop" in col) else 100
            style = f"left: {col_offsets[i]}px; min-width: {min_w}px;"
            # 店舗列はモバイル用に「ドコモショップ」を除いた短縮名も出力（PC非表示・モバイル表示）
            is_shop_col = ("店舗" in col or "Shop" in col) and "スタッフ" not in col
            if is_shop_col:
                short_v = str(v).replace("ドコモショップ", "") if v else v
                cell_content = f'<span class="shop-full">{v}</span><span class="shop-short">{short_v}</span>'
            else:
                cell_content = str(v) if v is not None else ''
            html.append(f'<td class="text-left sticky-col" style="{style}">{cell_content}</td>')
            
        for sc, _, _, _ in score_cols_ordered:
            v = row.get(sc, 0)
            try:
                html.append(f'<td class="text-right">{float(v):.1f}</td>')
            except:
                html.append(f'<td class="text-right">{v}</td>')
        
        total = row.get('Total_Score', 0)
        try: total_str = f'{float(total):.1f}'
        except: total_str = str(total)
        html.append(f'<td class="text-right font-bold">{total_str}</td>')
        
        if show_rank:
            rank = row.get('Rank', '')
            try: rank_str = f'{int(rank)}位'
            except: rank_str = str(rank)
            html.append(f'<td class="text-right">{rank_str}</td>')
            
        html.append('</tr>')
    html.append('</tbody></table></div>')
    return ''.join(html)

def count_page(df, target_month=None):
    st.header("件数")
    
    if df is None:
        st.info("データをアップロードしてください。")
        return

    # バックナンバー対応: target_monthが指定された月の設定を使用
    config = load_scoring_config(month=target_month)
    if not config or not config.get('items'):
         st.warning("評価設定がありません。設定を行ってください。")
         st.dataframe(df)
         return
         
    items = config.get('items', [])
    item_names = [item.get('name') for item in items if item.get('name')]
    
    # Optional: also merge with user master to ensure shop/team names are standardized
    user_master = load_user_master(month=target_month)
    if user_master.get('users'):
        df_master = pd.DataFrame(user_master['users'])
        name_col = 'スタッフ名' if 'スタッフ名' in df.columns else 'Name'
        shop_col = '店舗名' if '店舗名' in df.columns else 'Shop'
        team_col = 'チーム名' if 'チーム名' in df.columns else 'Team'
        
        if name_col in df.columns and 'name' in df_master.columns:
            merged_df = pd.merge(df, df_master[['name', 'shop', 'team']], 
                               left_on=name_col, right_on='name', how='left')
            if 'shop' in merged_df.columns:
                merged_df['店舗名'] = merged_df['shop'].fillna(merged_df[shop_col] if shop_col in merged_df.columns else pd.NA)
            elif shop_col in merged_df.columns:
                merged_df['店舗名'] = merged_df[shop_col]
            else:
                merged_df['店舗名'] = "未所属"
                
            if 'team' in merged_df.columns:
                merged_df['チーム名'] = merged_df['team'].fillna(merged_df[team_col] if team_col in merged_df.columns else pd.NA)
            elif team_col in merged_df.columns:
                merged_df['チーム名'] = merged_df[team_col]
            else:
                merged_df['チーム名'] = "未所属"
            
            # Use merged dataframe for calculations
            base_df = merged_df
        else:
            base_df = df.copy()
            if shop_col != '店舗名' and shop_col in base_df.columns: base_df['店舗名'] = base_df[shop_col]
            if team_col != 'チーム名' and team_col in base_df.columns: base_df['チーム名'] = base_df[team_col]
    else:
        base_df = df.copy()
        shop_col = '店舗' if '店舗' in base_df.columns else 'Shop'
        team_col = 'チーム' if 'チーム' in base_df.columns else 'Team'
        if shop_col != '店舗名' and shop_col in base_df.columns: base_df['店舗名'] = base_df[shop_col]
        if team_col != 'チーム名' and team_col in base_df.columns: base_df['チーム名'] = base_df[team_col]

    name_col = 'スタッフ名' if 'スタッフ名' in base_df.columns else 'Name'
    shop_col = '店舗名' if '店舗名' in base_df.columns else 'Shop'
    team_col = 'チーム名' if 'チーム名' in base_df.columns else 'Team'

    # 店舗の並び順を取得
    master_order = []
    if user_master.get('users'):
        unique_shops = []
        for u in user_master['users']:
            s = u.get('shop')
            if s and s not in unique_shops:
                unique_shops.append(s)
        master_order = unique_shops

    # 1. 個人件数表
    st.subheader("個人件数")
    show_personal = base_df.copy()
    
    # valid_items をここで定義
    valid_items = [i for i in item_names if i in show_personal.columns]

    if master_order:
        show_personal[shop_col] = pd.Categorical(show_personal[shop_col], categories=master_order, ordered=True)
        show_personal = show_personal.sort_values([shop_col, team_col, name_col])
    else:
        show_personal = show_personal.sort_values([shop_col, team_col, name_col])

    fixed = [(shop_col, '店舗'), (team_col, 'チーム'), (name_col, 'スタッフ名')]
    # 件数ページ用の項目列（配点などは0として表示）
    score_cols_p = []
    for iname in item_names:
        if iname in show_personal.columns:
            # 評価設定から評価軸と配点を取得
            item_cfg = next((it for it in items if it.get('name') == iname), {})
            etype = item_cfg.get('evaluation_type', 'relative')
            weight = item_cfg.get('weight', 0)
            
            # 実績値を numeric に変換
            show_personal[iname] = pd.to_numeric(show_personal[iname], errors='coerce').fillna(0)
            # 3段ヘッダーのために引数を渡す
            score_cols_p.append((iname, iname, etype, weight))

    # Total_Score カラムがないので一時的に作成（全項目の合計件数）
    show_personal['Total_Score'] = show_personal[valid_items].sum(axis=1)
    
    html_p = build_html_table(show_personal, fixed, score_cols_p, totals_label="合計件数", show_rank=False)
    st.markdown(html_p, unsafe_allow_html=True)

    st.markdown("<div style='height:30px'></div>", unsafe_allow_html=True)

    # 2. チーム件数表
    st.subheader("チーム件数")
    if team_col in base_df.columns:
        # リーダー情報の抽出: user_masterから直接「チーム→リーダー名」を決定（CSVの並び順に依存しない）
        def _norm_t(t):
            import unicodedata, re
            t = unicodedata.normalize('NFKC', str(t))
            return re.sub(r'[\s　]+', '', t)
        
        team_to_leader_map = {}
        if user_master.get('users'):
            for u in user_master['users']:
                tk = _norm_t(u.get('team', ''))
                if tk not in team_to_leader_map:
                    team_to_leader_map[tk] = u.get('name', '')
        
        agg_cols = {col: 'sum' for col in valid_items}
        team_counts = base_df.groupby([team_col, shop_col], as_index=False).agg(agg_cols)
        
        # user_masterから直接リーダーを付与
        team_counts['Leader'] = team_counts[team_col].apply(
            lambda t: team_to_leader_map.get(_norm_t(t), '')
        )
        team_sorted = team_counts
        if master_order:
            team_sorted[shop_col] = pd.Categorical(team_sorted[shop_col], categories=master_order, ordered=True)
            team_sorted = team_sorted.sort_values([shop_col, team_col])
        
        team_sorted['Total_Score'] = team_sorted[valid_items].sum(axis=1)
        fixed_t = [(shop_col, '店舗'), (team_col, 'チーム'), ('Leader', 'リーダー')]
        html_t = build_html_table(team_sorted, fixed_t, score_cols_p, totals_label="合計件数", show_rank=False, bg_color="#FFF5F5")
        st.markdown(html_t, unsafe_allow_html=True)
    else:
        st.info("チーム情報がありません")

    st.markdown("<div style='height:30px'></div>", unsafe_allow_html=True)

    # 3. 店舗件数表
    st.subheader("店舗件数")
    if shop_col in base_df.columns:
        # 店長情報の抽出（店舗別設定の shop_managers を使用）
        users = user_master.get('users', [])
        shop_managers = user_master.get('shop_managers', {})
        # shop_managers が空の場合は各店舗の先頭ユーザーにフォールバック
        if not shop_managers:
            for u in users:
                s = u.get('shop', '')
                if s and s not in shop_managers:
                    shop_managers[s] = u.get('name', '')

        agg_cols = {col: 'sum' for col in valid_items}
        shop_counts = base_df.groupby(shop_col, as_index=False).agg(agg_cols)
        
        shop_sorted = shop_counts.copy()
        shop_sorted['Manager'] = shop_sorted[shop_col].map(shop_managers).fillna('-')

        if master_order:
            shop_sorted[shop_col] = pd.Categorical(shop_sorted[shop_col], categories=master_order, ordered=True)
            shop_sorted = shop_sorted.sort_values(shop_col)
            
        shop_sorted['Total_Score'] = shop_sorted[valid_items].sum(axis=1)
        fixed_s = [(shop_col, '店舗'), ('Manager', '店長')]
        html_s = build_html_table(shop_sorted, fixed_s, score_cols_p, totals_label="合計件数", show_rank=False, bg_color="#F5FFF5")
        st.markdown(html_s, unsafe_allow_html=True)
    else:
        st.info("店舗情報がありません")


def comprehensive_page(df, target_month=None):
    st.header("総合")
    
    if df is None:
        st.info("データをアップロードしてください。")
        return

    # バックナンバー対応: target_monthが指定された月の設定を使用
    config = load_scoring_config(month=target_month)
    if not config or not config.get('items'):
        st.warning("評価設定がありません。設定を行ってください。")
        st.dataframe(df)
        return
    
    from utils.calculation import normalize_col_name as _ncn

    df_denom = get_denominator_df(df)
    scored_df = calculate_scores(df, config, df_denom)

    # ユーザーマスターとのマージ（バックナンバー対応: target_monthの月のマスターを使用）
    user_master = load_user_master(month=target_month)
    master_order = [] # 店舗の出現順序を保持
    if user_master.get('users'):
        df_master = pd.DataFrame(user_master['users'])
        # 店舗の並び順を取得（重複排除して出現順を維持）
        unique_shops = []
        for u in user_master['users']:
            s = u.get('shop')
            if s and s not in unique_shops:
                unique_shops.append(s)
        master_order = unique_shops

        name_col_s = 'スタッフ名'
        if name_col_s in scored_df.columns and 'name' in df_master.columns:
            merged = pd.merge(scored_df, df_master[['name', 'shop', 'team']],
                              left_on=name_col_s, right_on='name', how='left')
            if 'shop' in merged.columns:
                merged['店舗名'] = merged['shop'].fillna(merged.get('店舗名', ''))
            if 'team' in merged.columns:
                merged['チーム名'] = merged['team'].fillna(merged.get('チーム名', ''))
            scored_df = merged

    name_col = 'スタッフ名' if 'スタッフ名' in scored_df.columns else 'Name'
    shop_col = '店舗名' if '店舗名' in scored_df.columns else 'Shop'
    team_col = 'チーム名' if 'チーム名' in scored_df.columns else 'Team'

    items = config.get('items', [])

    # =========================================================
    # 個人成績テーブル
    # =========================================================
    st.subheader("個人成績")
    if 'Rank' not in scored_df.columns:
        st.error("ランキング計算エラー")
    else:
        # ソート: 設定順(店舗) → チーム → 順位
        ind_df = scored_df.copy()
        if master_order:
            # 店舗名にカテゴリ型を適用して、順番を維持したソートを可能にする
            ind_df[shop_col] = pd.Categorical(ind_df[shop_col], categories=master_order, ordered=True)
            ind_df = ind_df.sort_values([shop_col, team_col, 'Rank'])
        else:
            ind_df = ind_df.sort_values([shop_col, team_col, 'Rank'])

        fixed = [(shop_col, '店舗'), (team_col, 'チーム'), (name_col, 'スタッフ名')]
        score_cols_ord = [
            (sc, iname, item.get('evaluation_type', 'relative'), float(item.get('weight', 0)))
            for item in items
            for sc, iname in [(f"{_ncn(item.get('name',''))}_score", item.get('name',''))]
            if sc in ind_df.columns
        ]

        html = build_html_table(ind_df, fixed, score_cols_ord, items=items)
        st.markdown(html, unsafe_allow_html=True)

    st.markdown("<div style='height:30px'></div>", unsafe_allow_html=True)

    # =========================================================
    # チーム成績テーブル
    # =========================================================
    st.subheader("チーム成績")
    # リーダー判定を名簿優先にするため、原本も名簿順で正規化ソート
    raw_name_col = next((c for c in ['スタッフ名', '氏名', 'Name'] if c in df.columns), 'スタッフ名')
    df_for_agg = apply_robust_sorting(df, user_master, raw_name_col, '店舗名', 'チーム名')

    config['_user_master'] = user_master
    team_df = aggregate_team_scores(scored_df, df_original=df_for_agg, config=config)
    
    # 計算結果に関わらずLeaderをuser_masterから直接上書き（最終的な確実性を保証）
    if not team_df.empty and user_master.get('users'):
        import unicodedata as _ud, re as _re
        def _norm_ldr(t):
            t = _ud.normalize('NFKC', str(t))
            return _re.sub(r'[\s\u3000]+', '', t)
        team_ldr_map = {}
        for u in user_master['users']:
            tk = _norm_ldr(u.get('team', ''))
            if tk not in team_ldr_map:
                team_ldr_map[tk] = u.get('name', '')
        # org_aggのチーム名列を特定して上書き
        t_col_ldr = 'チーム名' if 'チーム名' in team_df.columns else ('チーム' if 'チーム' in team_df.columns else None)
        if t_col_ldr:
            team_df['Leader'] = team_df[t_col_ldr].apply(lambda t: team_ldr_map.get(_norm_ldr(t), ''))
    
    if not team_df.empty:
        t_col = 'チーム名' if 'チーム名' in team_df.columns else ('チーム' if 'チーム' in team_df.columns else 'Team')
        s_col = '店舗名' if '店舗名' in team_df.columns else ('店舗' if '店舗' in team_df.columns else 'Shop')
        leader_col = 'Leader' if 'Leader' in team_df.columns else None
        
        team_sorted = team_df.copy()
        if master_order:
            team_sorted[s_col] = pd.Categorical(team_sorted[s_col], categories=master_order, ordered=True)
            team_sorted = team_sorted.sort_values([s_col, 'Rank'])
        else:
            team_sorted = team_sorted.sort_values([s_col, 'Rank'])
        
        fixed_t = [(s_col, '店舗'), (t_col, 'チーム')]
        if leader_col:
            fixed_t.append((leader_col, 'リーダー'))

        team_score_cols = [
            (sc, iname, item.get('evaluation_type', 'relative'), float(item.get('weight', 0)))
            for item in items
            for sc, iname in [(f"{_ncn(item.get('name',''))}_score", item.get('name',''))]
            if sc in team_sorted.columns
        ]

        html_t = build_html_table(team_sorted, fixed_t, team_score_cols, bg_color="#FFF5F5", items=items)
        st.markdown(html_t, unsafe_allow_html=True)
    else:
        st.info("チーム情報がありません")

    st.markdown("<div style='height:30px'></div>", unsafe_allow_html=True)

    # =========================================================
    # 店舗成績テーブル
    # =========================================================
    st.subheader("店舗成績")
    config['_user_master'] = user_master
    shop_df = aggregate_shop_scores(scored_df, df_original=df, config=config)
    if not shop_df.empty:
        s_col2 = '店舗名' if '店舗名' in shop_df.columns else ('店舗' if '店舗' in shop_df.columns else 'Shop')
        mgr_col = 'Manager' if 'Manager' in shop_df.columns else None
        
        shop_sorted = shop_df.copy()
        if master_order:
            shop_sorted[s_col2] = pd.Categorical(shop_sorted[shop_sorted[s_col2].isin(master_order)][s_col2], categories=master_order, ordered=True)
            shop_sorted = shop_sorted.sort_values([s_col2, 'Rank'])
        else:
            shop_sorted = shop_sorted.sort_values([s_col2, 'Rank'])

        fixed_s = [(s_col2, '店舗')]
        if mgr_col:
            fixed_s.append((mgr_col, '店長'))

        shop_score_cols = [
            (sc, iname, item.get('evaluation_type', 'relative'), float(item.get('weight', 0)))
            for item in items
            for sc, iname in [(f"{_ncn(item.get('name',''))}_score", item.get('name',''))]
            if sc in shop_sorted.columns
        ]

        html_s = build_html_table(shop_sorted, fixed_s, shop_score_cols, bg_color="#F5FFF5", items=items)
        st.markdown(html_s, unsafe_allow_html=True)
    else:
        st.info("店舗情報がありません")


# =====================================================
# 成績詳細ページ
# =====================================================
def individual_detail_page(df, target_month=None):
    st.header("成績詳細")

    if df is None:
        st.info("データをアップロードしてください。")
        return

    # バックナンバー対応: target_monthが指定された月の設定を使用
    config = load_scoring_config(month=target_month)
    if not config or not config.get('items'):
        st.warning("評価設定がありません。管理者メニューから設定してください。")
        return

    df_denom = get_denominator_df(df)
    scored_df = calculate_scores(df, config, df_denom)

    # ユーザーマスターとマージ
    user_master = load_user_master(month=target_month)
    if user_master.get('users'):
        df_master = pd.DataFrame(user_master['users'])
        name_col = 'スタッフ名' if 'スタッフ名' in scored_df.columns else 'Name'
        if name_col in scored_df.columns and 'name' in df_master.columns:
            merged = pd.merge(scored_df, df_master[['name', 'shop', 'team']],
                              left_on=name_col, right_on='name', how='left')
            if 'shop' in merged.columns:
                merged['店舗名'] = merged['shop'].fillna(merged.get('店舗名', ''))
            if 'team' in merged.columns:
                merged['チーム名'] = merged['team'].fillna(merged.get('チーム名', ''))
            scored_df = merged

    name_col = 'スタッフ名' if 'スタッフ名' in scored_df.columns else 'Name'
    shop_col = '店舗名' if '店舗名' in scored_df.columns else 'Shop'
    team_col = 'チーム名' if 'チーム名' in scored_df.columns else 'Team'
    items = config.get('items', [])
    item_names = [item.get('name') for item in items if item.get('name')]

    # 表示単位の選択
    view_unit = st.radio("表示単位を選択", ["個人", "チーム", "店舗"], horizontal=True)
    
    if view_unit == "個人":
        target_df = scored_df.copy()
        target_name_col = name_col
        suffix = "名中"
        target_list = sorted(target_df[target_name_col].dropna().unique().tolist())
        selected_target = st.selectbox("スタッフを選択", target_list)
        denom_source = df_denom
    elif view_unit == "チーム":
        target_df = aggregate_team_scores(scored_df, df, config)
        target_name_col = 'チーム名'
        suffix = "チーム中"
        target_list = sorted(target_df[target_name_col].dropna().unique().tolist())
        selected_target = st.selectbox("チームを選択", target_list)
        denom_source = df_denom.copy() if df_denom is not None else pd.DataFrame()
        if not denom_source.empty:
            # マッピングを紐付け（チーム名が存在しないエラーを修正）
            if 'df_master' in locals() and 'name' in df_master.columns:
                denom_source = pd.merge(denom_source, df_master[['name', 'team']], left_on=name_col, right_on='name', how='left')
                denom_source['チーム名'] = denom_source['team'].fillna('-')
            if 'チーム名' in denom_source.columns:
                denom_source = denom_source.groupby('チーム名').sum(numeric_only=True).reset_index()
    else: # 店舗
        target_df = aggregate_shop_scores(scored_df, df, config)
        target_name_col = '店舗名'
        suffix = "店舗中"
        target_list = sorted(target_df[target_name_col].dropna().unique().tolist())
        selected_target = st.selectbox("店舗を選択", target_list)
        denom_source = df_denom.copy() if df_denom is not None else pd.DataFrame()
        if not denom_source.empty:
            # マッピングを紐付け（店舗名が存在しないエラーを修正）
            if 'df_master' in locals() and 'name' in df_master.columns:
                denom_source = pd.merge(denom_source, df_master[['name', 'shop']], left_on=name_col, right_on='name', how='left')
                denom_source['店舗名'] = denom_source['shop'].fillna('-')
            if '店舗名' in denom_source.columns:
                denom_source = denom_source.groupby('店舗名').sum(numeric_only=True).reset_index()

    if not selected_target:
        return

    row = target_df[target_df[target_name_col] == selected_target].iloc[0]

    # プロフィールカード（詳細表示版）
    total_count = len(target_df)
    total_weight = sum(float(it.get('weight', 0)) for it in items)
    
    col1, col2 = st.columns(2)
    with col1:
        st.metric("総合順位", f"{int(row.get('Rank', 0))}位 / {total_count}{suffix}")
    with col2:
        st.metric("総合スコア", f"{row.get('Total_Score', 0):.1f}点 / {int(total_weight)}点満点中")
    
    if view_unit == "個人":
        c_shop, c_team = st.columns(2)
        with c_shop:
            st.markdown(f"**店舗**")
            st.markdown(f"<div style='background-color:#f8f9fa; padding:10px; border-radius:5px; border-left:5px solid #1A73E8; font-size:16px;'>{row.get(shop_col, '-')}</div>", unsafe_allow_html=True)
        with c_team:
            st.markdown(f"**チーム**")
            st.markdown(f"<div style='background-color:#f8f9fa; padding:10px; border-radius:5px; border-left:5px solid #34A853; font-size:16px;'>{row.get(team_col, '-')}</div>", unsafe_allow_html=True)
    elif view_unit == "チーム":
        st.markdown(f"**所属店舗**")
        st.markdown(f"<div style='background-color:#f8f9fa; padding:10px; border-radius:5px; border-left:5px solid #1A73E8; font-size:16px;'>{row.get(shop_col, '-')}</div>", unsafe_allow_html=True)

    st.markdown("---")
    
    # --- レーダーチャートの描画 ---
    chart_data = []
    from utils.calculation import normalize_col_name as _ncn
    for item in items:
        iname = item.get('name', '')
        weight = item.get('weight', 0)
        norm_name = _ncn(iname)
        score_col = f"{norm_name}_score"
        score_val = row.get(score_col, 0)
        achievement = (score_val / weight * 100) if weight > 0 else 0
        chart_data.append({"item": iname, "value": achievement})
    
    if chart_data:
        chart_df = pd.DataFrame(chart_data)
        fig = go.Figure()
        fig.add_trace(go.Scatterpolar(
            r=chart_df['value'].tolist() + [chart_df['value'].iloc[0]],
            theta=chart_df['item'].tolist() + [chart_df['item'].iloc[0]],
            fill='toself',
            name=selected_target,
            line=dict(color='#1A73E8'),
            fillcolor='rgba(26, 115, 232, 0.2)'
        ))

        fig.update_layout(
            polar=dict(
                radialaxis=dict(visible=True, range=[0, 100], tickfont=dict(size=10), gridcolor="#eee"),
                angularaxis=dict(tickfont=dict(size=11), rotation=90, direction="clockwise"),
                bgcolor="white"
            ),
            showlegend=False,
            margin=dict(l=40, r=40, t=20, b=20),
            height=500,
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
        )
        st.subheader(f"{view_unit}バランスマップ（達成率％）")
        st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})

    st.subheader("項目別スコア詳細")
    
    # 項目別スコアテーブル
    score_details = []
    etype_jp = {
        'relative': '相対評価', 'absolute': '絶対評価',
        'relative_absolute': '相対絶対', 
        'team_relative': 'チーム相対', 'team_absolute': 'チーム絶対', 
        'shop_achievement': '店舗達成率', 'individual_achievement': '個人達成'
    }
    
    for item in items:
        iname = item.get('name', '')
        weight = item.get('weight', 0)
        norm_name = _ncn(iname)
        score_col = f"{norm_name}_score"
        
        # 実績値とスコアの取得
        perf_val = row.get(iname, row.get(norm_name, 0))
        score_val = row.get(score_col, 0)
        achievement = (score_val / weight * 100) if weight > 0 else 0
        
        etype_display = etype_jp.get(item.get('evaluation_type'), item.get('evaluation_type', '-'))
        
        score_details.append({
            "項目名": iname,
            "実績": round(perf_val, 1) if not pd.isna(perf_val) else 0,
            "獲得スコア": round(score_val, 1),
            "配点": weight,
            "達成率": f"{achievement:.1f}%",
            "評価軸": etype_display
        })
    
    st.dataframe(pd.DataFrame(score_details), hide_index=True, use_container_width=True)

    # --- 成績分母の内訳 ---
    if not denom_source.empty:
        denom_row = denom_source[denom_source[target_name_col] == selected_target]
        if not denom_row.empty:
            st.markdown("---")
            st.subheader(f"🔢 {view_unit}成績分母の内訳")
            denom_disp_row = denom_row.copy()
            for col in denom_disp_row.columns:
                if pd.api.types.is_numeric_dtype(denom_disp_row[col]):
                    denom_disp_row[col] = denom_disp_row[col].round(1)
            
            display_cols = [
                target_name_col, '基準勤務時間', '端末販売数(合計)', '端末販売数出勤日', 
                'オーダー数(合計)', '分母内訳_端末販売数', '分母内訳_オーダー数', '成績分母'
            ]
            available_cols = [c for c in display_cols if c in denom_disp_row.columns]
            denom_disp = denom_disp_row[available_cols].T.reset_index()
            denom_disp.columns = ["項目", "値"]
            st.dataframe(denom_disp, hide_index=True, use_container_width=True)


# =====================================================
# 分析ページ
# =====================================================
def analysis_page(df):
    st.header("分析")

    if df is None:
        st.info("データをアップロードしてください。")
        return

    config = load_scoring_config()
    if not config or not config.get('items'):
        st.warning("評価設定がありません。管理者メニューから設定してください。")
        return

    df_denom = get_denominator_df(df)
    scored_df = calculate_scores(df, config, df_denom)

    # ユーザーマスターとマージ
    user_master = load_user_master()
    if user_master.get('users'):
        df_master = pd.DataFrame(user_master['users'])
        name_col_s = 'スタッフ名' if 'スタッフ名' in scored_df.columns else 'Name'
        if name_col_s in scored_df.columns and 'name' in df_master.columns:
            merged = pd.merge(scored_df, df_master[['name', 'shop', 'team']],
                              left_on=name_col_s, right_on='name', how='left')
            if 'shop' in merged.columns:
                merged['店舗名'] = merged['shop'].fillna(merged.get('店舗名', ''))
            if 'team' in merged.columns:
                merged['チーム名'] = merged['team'].fillna(merged.get('チーム名', ''))
            scored_df = merged

    name_col = 'スタッフ名' if 'スタッフ名' in scored_df.columns else 'Name'
    shop_col = '店舗名' if '店舗名' in scored_df.columns else 'Shop'
    team_col = 'チーム名' if 'チーム名' in scored_df.columns else 'Team'
    items = config.get('items', [])
    item_names = [item.get('name') for item in items if item.get('name')]

    st.subheader("比較・シミュレーション")
    
    comp_unit = st.radio("比較単位", ["個人", "チーム", "店舗"], horizontal=True, key="comp_unit")
    
    # 対象データの準備
    if comp_unit == "個人":
        target_df = scored_df.copy()
        target_name_col = name_col
        col_a_label, col_b_label = "スタッフA", "スタッフB"
    elif comp_unit == "チーム":
        target_df = aggregate_team_scores(scored_df, df, config)
        target_name_col = 'チーム名'
        col_a_label, col_b_label = "チームA", "チームB"
    else:
        target_df = aggregate_shop_scores(scored_df, df, config)
        target_name_col = '店舗名'
        col_a_label, col_b_label = "店舗A", "店舗B"

    if target_df.empty:
        st.warning("比較対象となるデータがありません。")
    else:
        target_list = sorted(target_df[target_name_col].dropna().unique().tolist())
        
        col_sel_a, col_sel_b = st.columns(2)
        with col_sel_a:
            target_a = st.selectbox(f"{col_a_label}を選択", target_list, key="sel_a")
        with col_sel_b:
            target_b = st.selectbox(f"{col_b_label}を選択", target_list, key="sel_b", index=min(1, len(target_list)-1))

        if target_a and target_b:
            row_a = target_df[target_df[target_name_col] == target_a].iloc[0]
            row_b = target_df[target_df[target_name_col] == target_b].iloc[0]
            
            # --- 1. 比較サマリー ---
            c1, c2 = st.columns(2)
            with c1:
                st.markdown(f"#### {target_a}")
                st.metric("順位", f"{int(row_a['Rank'])}位", f"{int(row_b['Rank'] - row_a['Rank'])}位差" if row_a['Rank'] != row_b['Rank'] else None, delta_color="inverse")
                st.metric("合計スコア", f"{row_a['Total_Score']:.1f}点", f"{row_a['Total_Score'] - row_b['Total_Score']:.1f}点差")
            with c2:
                st.markdown(f"#### {target_b}")
                st.metric("順位", f"{int(row_b['Rank'])}位")
                st.metric("合計スコア", f"{row_b['Total_Score']:.1f}点")

            # --- 2. 項目別比較テーブル ---
            st.markdown("##### 項目別実績・スコア比較")
            comp_data = []
            from utils.calculation import normalize_col_name as _ncn
            for item in items:
                iname = item.get('name', '')
                norm_name = _ncn(iname)
                score_col = f"{norm_name}_score"
                
                perf_a = row_a.get(iname, row_a.get(norm_name, 0))
                score_a = row_a.get(score_col, 0)
                perf_b = row_b.get(iname, row_b.get(norm_name, 0))
                score_b = row_b.get(score_col, 0)
                
                comp_data.append({
                    "項目": iname,
                    f"{target_a}実績": f"{perf_a:.1f}",
                    f"{target_a}スコア": f"{score_a:.1f}",
                    f"{target_b}実績": f"{perf_b:.1f}",
                    f"{target_b}スコア": f"{score_b:.1f}",
                    "スコア差": f"{score_a - score_b:+.1f}"
                })
            st.table(pd.DataFrame(comp_data))

            # --- 3. シミュレーション ---
            st.markdown("---")
            st.subheader(f"🧪 {target_a} の獲得数シミュレーション")
            st.write(f"「もし {target_a} があと●件多く獲得していたら？」をシミュレーションします。")
            
            sim_inputs = {}
            valid_items = [it for it in items if it.get('name') in df.columns or _ncn(it.get('name')) in df.columns]
            
            # 入力UI
            cols = st.columns(min(len(valid_items), 3) if valid_items else 1)
            for i, item in enumerate(valid_items):
                iname = item.get('name')
                with cols[i % len(cols)]:
                    sim_inputs[iname] = st.number_input(f"{iname} を追加", min_value=-100, max_value=100, value=0, key=f"sim_in_{iname}")
            
            if st.button("シミュレーションを実行", type="primary"):
                # シミュレーション用データの作成
                sim_df = df.copy()
                if comp_unit == "個人":
                    # 対象者の実績を加算
                    for iname, add_val in sim_inputs.items():
                        if add_val != 0:
                            sim_df.loc[sim_df[name_col] == target_a, iname] += add_val
                elif comp_unit == "チーム":
                    # チーム全員に均等に割り振る（または代表者に加算。ここでは合計が合えば良いので代表者に加算）
                    team_members = df[df['チーム名'] == target_a][name_col].tolist()
                    if team_members:
                        first_member = team_members[0]
                        for iname, add_val in sim_inputs.items():
                            sim_df.loc[sim_df[name_col] == first_member, iname] += add_val
                else: # 店舗
                    shop_members = df[df['店舗名'] == target_a][name_col].tolist()
                    if shop_members:
                        first_member = shop_members[0]
                        for iname, add_val in sim_inputs.items():
                            sim_df.loc[sim_df[name_col] == first_member, iname] += add_val

                # 再計算（既存ロジック calculate_scores を使用）
                sim_scored = calculate_scores(sim_df, config, df_denom)
                
                # 組織集計が必要な場合
                if comp_unit == "チーム":
                    sim_result_df = aggregate_team_scores(sim_scored, sim_df, config)
                elif comp_unit == "店舗":
                    sim_result_df = aggregate_shop_scores(sim_scored, sim_df, config)
                else:
                    sim_result_df = sim_scored
                
                # 結果の抽出
                sim_row_a = sim_result_df[sim_result_df[target_name_col] == target_a].iloc[0]
                
                # 表示
                st.markdown("#### 🏁 シミュレーション結果")
                
                rank_diff = int(row_a['Rank'] - sim_row_a['Rank'])
                score_diff = sim_row_a['Total_Score'] - row_a['Total_Score']
                
                res_col1, res_col2 = st.columns(2)
                with res_col1:
                    st.metric("シミュレーション後の順位", f"{int(sim_row_a['Rank'])}位", f"{rank_diff}位アップ" if rank_diff > 0 else (f"{abs(rank_diff)}位ダウン" if rank_diff < 0 else None))
                with res_col2:
                    st.metric("シミュレーション後のスコア", f"{sim_row_a['Total_Score']:.1f}点", f"+{score_diff:.1f}点")
                
                # テキストメッセージ
                if rank_diff > 0:
                    st.success(f"🎊 あとこれだけ獲得していれば、**{int(sim_row_a['Rank'])}位** になれました！")
                elif score_diff > 0:
                    st.info(f"得点は **{sim_row_a['Total_Score']:.1f}点** にアップしますが、順位は変わりませんでした。")
                else:
                    st.warning("実績に変化がありません。")

                # アイテム別の影響
                st.markdown("##### 項目別スコアへの影響")
                impact_data = []
                for item in items:
                    iname = item.get('name')
                    norm = _ncn(iname)
                    sc_col = f"{norm}_score"
                    old_sc = row_a.get(sc_col, 0)
                    new_sc = sim_row_a.get(sc_col, 0)
                    if sim_inputs.get(iname, 0) != 0 or old_sc != new_sc:
                        impact_data.append({
                            "項目": iname,
                            "元実績": f"{float(row_a.get(iname, 0)):.1f}",
                            "修正後実績": f"{float(sim_row_a.get(iname, 0)):.1f}",
                            "元スコア": f"{float(old_sc):.1f}",
                            "修正後スコア": f"{float(new_sc):.1f}",
                            "得点増分": f"{float(new_sc - old_sc):+.1f}"
                        })
                if impact_data:
                    st.table(pd.DataFrame(impact_data))


# =====================================================
# 配点シミュレーションページ
# =====================================================
def simulation_page(df):
    st.header("配点シミュレーション")
    st.markdown("評価設定を仮で変更して、スコアへの影響をシミュレーションします。実際の設定は変更されません。")

    if df is None:
        st.info("データをアップロードしてください。")
        return

    config = load_scoring_config()
    if not config or not config.get('items'):
        st.warning("評価設定がありません。管理者メニューから設定してください。")
        return

    items = config.get('items', [])
    item_names = [item.get('name') for item in items]

    st.subheader("配点の仮変更")
    st.markdown("各項目の配点を調整して「シミュレーション実行」を押してください。")

    import copy
    sim_config = copy.deepcopy(config)
    sim_items = sim_config.get('items', [])

    cols_per_row = 3
    sim_weights = {}

    rows = [item_names[i:i+cols_per_row] for i in range(0, len(item_names), cols_per_row)]
    for row_items in rows:
        cols = st.columns(cols_per_row)
        for j, iname in enumerate(row_items):
            item = next((it for it in sim_items if it.get('name') == iname), {})
            current_w = item.get('weight', 0)
            new_w = cols[j].number_input(
                f"{iname} 配点",
                min_value=0, max_value=500,
                value=int(current_w),
                step=5,
                key=f"sim_{iname}"
            )
            sim_weights[iname] = new_w

    # 合計配点表示
    total_sim = sum(sim_weights.values())
    st.info(f"シミュレーション合計配点: **{total_sim}点**")

    if st.button("▶ シミュレーション実行", type="primary"):
        # sim_config に新配点を適用
        for item in sim_items:
            iname = item.get('name')
            if iname in sim_weights:
                item['weight'] = sim_weights[iname]
                if 'weights' in item:
                    item['weights']['w1'] = sim_weights[iname]

        df_denom = get_denominator_df(df)
        sim_scored = calculate_scores(df, sim_config, df_denom)

        name_col = 'スタッフ名' if 'スタッフ名' in sim_scored.columns else 'Name'

        st.subheader("シミュレーション結果")
        sim_disp_cols = ['Rank', name_col, 'Total_Score'] + [
            f"{n}_score" for n in item_names if f"{n}_score" in sim_scored.columns
        ]
        valid_cols = [c for c in sim_disp_cols if c in sim_scored.columns]
        show = sim_scored.sort_values('Rank')[valid_cols].copy()
        show['Total_Score'] = show['Total_Score'].round(1)
        st.dataframe(show, hide_index=True, use_container_width=True)

        # 現行vs シミュレーション比較
        st.subheader("現行 vs シミュレーション 比較")
        orig_scored = calculate_scores(df, config, df_denom)
        compare = pd.merge(
            orig_scored[[name_col, 'Total_Score', 'Rank']].rename(
                columns={'Total_Score': '現行スコア', 'Rank': '現行順位'}
            ),
            sim_scored[[name_col, 'Total_Score', 'Rank']].rename(
                columns={'Total_Score': 'SIMスコア', 'Rank': 'SIM順位'}
            ),
            on=name_col, how='outer'
        )
        compare['スコア差'] = (compare['SIMスコア'] - compare['現行スコア']).round(1)
        compare['順位差'] = (compare['現行順位'] - compare['SIM順位']).astype(int)
        compare = compare.sort_values('SIM順位')
        st.dataframe(compare, hide_index=True, use_container_width=True)


# =====================================================
# ルール説明ページ
# =====================================================
def rules_page(config=None):
    import pandas as pd
    st.header("📖 \u30eb\u30fc\u30eb\u8aac\u660e")
    st.markdown("\u3053\u306e\u30da\u30fc\u30b8\u3067\u306f\u300c\u30d6\u30e9\u30a4\u30c8\u30d1\u30b9\u304f\u3093\u300d\u306e\u6210\u7e3e\u8a55\u4fa1\u306e\u4ed5\u7d44\u307f\u3092\u8aac\u660e\u3057\u307e\u3059\u3002")

    # --- 1. \u30b7\u30b9\u30c6\u30e0\u6982\u8981 ---
    st.markdown("---")
    st.subheader("🏗️ \u3053\u306e\u30b7\u30b9\u30c6\u30e0\u306b\u3064\u3044\u3066")
    st.info(
        "\u30d6\u30e9\u30a4\u30c8\u30d1\u30b9\u304f\u3093\u306f\u3001\u6708\u6b21\u306e\u6210\u7e3e\u30c7\u30fc\u30bf\u3092\u3082\u3068\u306b **\u500b\u4eba\u30fb\u30c1\u30fc\u30e0\u30fb\u5e97\u8217** \u305d\u308c\u305e\u308c\u306e\u30b9\u30b3\u30a2\u3092\u81ea\u52d5\u8a08\u7b97\u3057\u3001"
        "\u30e9\u30f3\u30ad\u30f3\u30b0\u3092\u53ef\u8996\u5316\u3059\u308b\u6210\u7e3e\u7ba1\u7406\u30b7\u30b9\u30c6\u30e0\u3067\u3059\u3002\n\n"
        "\u6bce\u6708\u306e\u8a55\u4fa1\u9805\u76ee\u30fb\u914d\u70b9\u306f\u7ba1\u7406\u8005\u304c\u8a2d\u5b9a\u3057\u3066\u304a\u308a\u3001\u5404\u30b9\u30bf\u30c3\u30d5\u306e\u5b9f\u7e3e\u30c7\u30fc\u30bf\u306b\u57fa\u3065\u3044\u3066\u81ea\u52d5\u7684\u306b\u7dcf\u5408\u30b9\u30b3\u30a2\u304c\u8a08\u7b97\u3055\u308c\u307e\u3059\u3002"
    )
    col1, col2, col3 = st.columns(3)
    items = config.get('items', []) if config else []
    total_weight = sum(float(i.get('weight', 0)) for i in items)
    with col1:
        st.metric("\u8a55\u4fa1\u5358\u4f4d", "\u500b\u4eba / \u30c1\u30fc\u30e0 / \u5e97\u8217")
    with col2:
        st.metric("\u73fe\u5728\u306e\u6e80\u70b9", f"{total_weight:.0f} \u70b9")
    with col3:
        st.metric("\u8a55\u4fa1\u8ef8\u306e\u7a2e\u985e", "5\u7a2e\u985e")

    # --- 2. \u8a55\u4fa1\u8ef8\u306e\u8aac\u660e ---
    st.markdown("---")
    st.subheader("\u2696\ufe0f 5\u7a2e\u985e\u306e\u8a55\u4fa1\u8ef8")
    st.markdown("\u5404\u8a55\u4fa1\u9805\u76ee\u306b\u306f\u300c\u8a55\u4fa1\u8ef8\u300d\u304c\u8a2d\u5b9a\u3055\u308c\u3066\u304a\u308a\u3001\u30b9\u30b3\u30a2\u306e\u8a08\u7b97\u65b9\u6cd5\u304c\u7570\u306a\u308a\u307e\u3059\u3002")

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "\u2460 \u76f8\u5bfe\u8a55\u4fa1", "\u2461 \u7d76\u5bfe\u8a55\u4fa1", "\u2462 \u76f8\u5bfe\u7d76\u5bfe", "\u2463 \u5e97\u8217\u9054\u6210\u7387", "\u2464 \u500b\u4eba\u9054\u6210"
    ])

    with tab1:
        st.markdown("### \u76f8\u5bfe\u8a55\u4fa1 `relative`")
        st.markdown(
            "**\u30c1\u30fc\u30e0\u5185\u306e\u9806\u4f4d**\u306b\u5fdc\u3058\u3066\u30b9\u30b3\u30a2\u304c\u5909\u308f\u308a\u307e\u3059\u30021\u4f4d\u304c\u6e80\u70b9\u3001\u4ee5\u964d\u306f\u7b49\u5206\u5272\u3067\u6e1b\u70b9\u3055\u308c\u307e\u3059\u3002\n\n"
            "- \u5b9f\u7e3e\u304c **0\u4ef6** \u306e\u5834\u5408\u306f\u9806\u4f4d\u306b\u95a2\u308f\u3089\u305a **0\u70b9**\n"
            "- 1\u30b9\u30c6\u30c3\u30d7 = \u914d\u70b9 \u00f7 \u53c2\u52a0\u4eba\u6570"
        )
        st.markdown("#### \u8a08\u7b97\u4f8b\uff08\u914d\u70b970\u70b9\u30fb10\u4eba\u53c2\u52a0\uff09")
        st.dataframe(pd.DataFrame({
            "\u9806\u4f4d": ["1\u4f4d", "2\u4f4d", "3\u4f4d", "\u2026", "10\u4f4d"],
            "\u30b9\u30b3\u30a2": ["70.0\u70b9", "63.0\u70b9", "56.0\u70b9", "\u2026", "7.0\u70b9"],
            "\u8a08\u7b97": ["70 - 0\u00d77", "70 - 1\u00d77", "70 - 2\u00d77", "\u2026", "70 - 9\u00d77"]
        }), use_container_width=True, hide_index=True)

    with tab2:
        st.markdown("### \u7d76\u5bfe\u8a55\u4fa1 `absolute`")
        st.markdown(
            "\u8a2d\u5b9a\u3055\u308c\u305f**\u76ee\u6a19\u9054\u6210\u7387**\u306b\u5fdc\u3058\u3066\u3001Lv1/Lv2/Lv3 \u306e\u6bb5\u968e\u7684\u306a\u914d\u70b9\u304c\u4ed8\u4e0e\u3055\u308c\u307e\u3059\u3002\n\n"
            "- **\u9054\u6210\u7387** = `(\u5b9f\u7e3e \u00f7 \u5206\u6bcd) \u00f7 \u76ee\u6a19`\n"
            "- Lv1\u3057\u304d\u3044\u5024\u4ee5\u4e0a \u2192 Lv1\u914d\u70b9\uff08\u6700\u9ad8\uff09\n"
            "- Lv2\u3057\u304d\u3044\u5024\u4ee5\u4e0a \u2192 Lv2\u914d\u70b9\n"
            "- \u672a\u9054 \u2192 0\u70b9"
        )
        st.markdown("#### \u8a08\u7b97\u4f8b\uff08Lv1=80%\u679c\u305f\u305950\u70b9\u3001Lv2=70%\u679c\u305f\u305940\u70b9\uff09")
        st.dataframe(pd.DataFrame({
            "\u9054\u6210\u7387": ["80%\u4ee5\u4e0a", "70%\u4ee5\u4e0a", "70%\u672a\u6e80"],
            "\u30b9\u30b3\u30a2": ["50\u70b9 (Lv1)", "40\u70b9 (Lv2)", "0\u70b9"],
        }), use_container_width=True, hide_index=True)

    with tab3:
        st.markdown("### \u76f8\u5bfe\u7d76\u5bfe\u8a55\u4fa1 `relative_absolute`")
        st.markdown(
            "**\u9054\u6210\u7387\u306e\u30e9\u30f3\u30ad\u30f3\u30b0**\uff08\u76f8\u5bfe\u8a55\u4fa1\uff09\u306b\u52a0\u3048\u3001\u5168\u4f53\u5e73\u5747\u306e50%\u3092\u30af\u30ea\u30a2\u3057\u3066\u3044\u308c\u3070**\u6700\u4f4e\u4fdd\u8a3c**\u304c\u3064\u304f\u65b9\u5f0f\u3067\u3059\u3002\n\n"
            "1. \u9054\u6210\u7387\u3067\u5168\u54e1\u3092\u30e9\u30f3\u30ad\u30f3\u30b0\u3057\u3001\u7b49\u5206\u5272\u6e1b\u70b9\u3067\u30b9\u30b3\u30a2\u3092\u7b97\u51fa\n"
            "2. \u81ea\u5206\u306e\u9054\u6210\u7387\u304c\u300c\u5168\u4f53\u5e73\u5747\u306e50%\u300d\u3092\u8d85\u3048\u3066\u3044\u308c\u3070\u3001\u914d\u70b9\u306e **50%** \u3092\u6700\u4f4e\u4fdd\u8a3c\u3068\u3057\u3066\u4ed8\u4e0e"
        )
        st.info("\u2605 \u9811\u5f35\u3063\u3066\u3044\u308b\u304c\u9806\u4f4d\u304c\u4f4e\u3044\u30b9\u30bf\u30c3\u30d5\u306e\u30b9\u30b3\u30a2\u304c\u5e95\u4e0a\u3052\u3055\u308c\u308b\u4ed5\u7d44\u307f\u3067\u3059\u3002")

    with tab4:
        st.markdown("### \u5e97\u8217\u9054\u6210\u7387 `shop_achievement`")
        st.markdown(
            "**\u5e97\u8217\u5168\u4f53\u306e\u5b9f\u7e3e\u5408\u8a08**\u304c\u76ee\u6a19\u3092\u30af\u30ea\u30a2\u3057\u305f\u304b\u3069\u3046\u304b\u3067\u5224\u5b9a\u3057\u307e\u3059\u3002\n"
            "\u5e97\u8217\u3067\u9054\u6210\u3057\u3066\u3044\u308c\u3070\u3001\u6240\u5c5e\u3059\u308b\u5168\u30b9\u30bf\u30c3\u30d5\u306b**\u4e00\u5f8b\u3067\u540c\u3058\u70b9\u6570**\u304c\u4ed8\u4e0e\u3055\u308c\u307e\u3059\u3002\n\n"
            "- **\u9054\u6210\u7387** = `\u5e97\u8217\u5408\u8a08\u5b9f\u7e3e \u00f7 \u5e97\u8217\u76ee\u6a19`\n"
            "- 100%\u9054\u6210 \u2192 \u5168\u54e1\u6e80\u70b9\u3001\u672a\u9054 \u2192 \u5168\u54e1 0\u70b9"
        )
        st.info("\u2605 \u5e97\u8217\u4e00\u4e38\u3068\u306a\u3063\u3066\u76ee\u6a19\u9054\u6210\u3092\u76ee\u6307\u3059\u8a55\u4fa1\u65b9\u5f0f\u3067\u3059\u3002")

    with tab5:
        st.markdown("### \u500b\u4eba\u9054\u6210 `individual_achievement`")
        st.markdown(
            "**\u81ea\u5206\u304c\u500b\u4eba\u76ee\u6a19\u3092\u9054\u6210\u3057\u305f\u304b\u3069\u3046\u304b**\u3060\u3051\u30670\u304b\u6e80\u70b9\u304c\u6c7a\u307e\u308a\u307e\u3059\u3002\u5206\u6bcd\u306f\u4f7f\u3044\u307e\u305b\u3093\u3002"
        )
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("#### \u500b\u4eba\u306e\u5224\u5b9a\uff08\u81ea\u5df1\u5b8c\u7d50\uff09")
            st.dataframe(pd.DataFrame({
                "\u6761\u4ef6": ["\u5b9f\u7e3e \u2265 \u500b\u4eba\u76ee\u6a19", "\u5b9f\u7e3e < \u500b\u4eba\u76ee\u6a19"],
                "\u30b9\u30b3\u30a2": ["\u6e80\u70b9", "0\u70b9"],
            }), use_container_width=True, hide_index=True)
            st.caption("\u203b \u30c1\u30fc\u30e0\u304c\u9054\u6210\u3057\u3066\u3044\u3066\u3082\u81ea\u5206\u304c\u672a\u9054\u306a\u30890\u70b9\u306e\u307e\u307e")
        with c2:
            st.markdown("#### \u30c1\u30fc\u30e0\u30fb\u5e97\u8217\u306e\u5224\u5b9a")
            st.dataframe(pd.DataFrame({
                "\u6761\u4ef6": ["\u5408\u8a08 \u2265 \u500b\u4eba\u76ee\u6a19\u00d7\u4eba\u6570", "\u5408\u8a08 < \u500b\u4eba\u76ee\u6a19\u00d7\u4eba\u6570"],
                "\u30b9\u30b3\u30a2": ["\u6e80\u70b9", "0\u70b9"],
            }), use_container_width=True, hide_index=True)
            st.caption("\u203b 1\u4eba\u304c\u591a\u304f\u58f2\u308c\u3070\u7d44\u7e54\u5168\u4f53\u3092\u30ab\u30d0\u30fc\u3067\u304d\u308b")
        st.markdown("#### \u5177\u4f53\u4f8b\uff08\u500b\u4eba\u76ee\u6a19=1\u4ef6\u3001A\u30c1\u30fc\u30e03\u540d\uff09")
        st.dataframe(pd.DataFrame({
            "\u5358\u4f4d": ["A\u3055\u3093\uff082\u4ef6\uff09", "B\u3055\u3093\uff081\u4ef6\uff09", "C\u3055\u3093\uff080\u4ef6\uff09", "A\u30c1\u30fc\u30e0\u5408\u8a08\uff083\u4ef6\uff09"],
            "\u5224\u5b9a": ["2 \u2265 1 \u2705", "1 \u2265 1 \u2705", "0 < 1 \u274c", "3 \u2265 (1\u00d73=3) \u2705"],
            "\u30b9\u30b3\u30a2": ["10\u70b9", "10\u70b9", "0\u70b9\uff08\u30c1\u30fc\u30e0\u9054\u6210\u3067\u3082\uff09", "10\u70b9"],
        }), use_container_width=True, hide_index=True)

    # --- 3. 分母（SAITO式）の説明 ---
    st.markdown("---")
    st.subheader("🔢 分母（SAITO式）とは")
    st.markdown(
        "**相対評価・相対絶対評価**では、単純な件数ではなく「稼働量に対してどれだけ貢献したか」で比較します。\n\n"
        "オーダー数や端末販売数など、稼働状況を総合的に加味して算出した **公正な基準値** です。"
        "単純な出勤日数だけでなく、実際の活動量が反映されます。"
    )

    # --- 4. \u73fe\u5728\u306e\u8a55\u4fa1\u9805\u76ee\u4e00\u89a7 ---
    if items:
        st.markdown("---")
        st.subheader("📋 \u73fe\u5728\u306e\u8a55\u4fa1\u9805\u76ee\u4e00\u89a7")
        eval_type_map = {
            'relative': '\u76f8\u5bfe\u8a55\u4fa1',
            'absolute': '\u7d76\u5bfe\u8a55\u4fa1',
            'relative_absolute': '\u76f8\u5bfe\u7d76\u5bfe',
            'shop_achievement': '\u5e97\u8217\u9054\u6210\u7387',
            'individual_achievement': '\u500b\u4eba\u9054\u6210',
            'team_absolute': '\u30c1\u30fc\u30e0\u7d76\u5bfe',
            'team_relative': '\u30c1\u30fc\u30e0\u76f8\u5bfe',
        }
        rows = [{
            "#": i + 1,
            "\u9805\u76ee\u540d": item.get('name', ''),
            "\u914d\u70b9": f"{item.get('weight', 0)}\u70b9",
            "\u8a55\u4fa1\u8ef8": eval_type_map.get(item.get('evaluation_type', ''), item.get('evaluation_type', '')),
            "\u5206\u6bcd\u7a2e\u5225": item.get('denominator_type', 'SAITO\u5f0f'),
        } for i, item in enumerate(items)]
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    # --- 5. \u5404\u30da\u30fc\u30b8\u306e\u898b\u65b9 ---
    st.markdown("---")
    st.subheader("📊 \u5404\u30da\u30fc\u30b8\u306e\u898b\u65b9")
    for page, desc in {
        "📊 \u7dcf\u5408": "\u5168\u8a55\u4fa1\u9805\u76ee\u306e\u30b9\u30b3\u30a2\u3092\u5408\u7b97\u3057\u305f\u6210\u7e3e\u8868\u3002\u540d\u7c3f\u9806\u306b\u8868\u793a\u3002\u500b\u4eba\u30fb\u30c1\u30fc\u30e0\u30fb\u5e97\u8217\u306e3\u6bb5\u69cb\u6210\u3002",
        "🔢 \u4ef6\u6570": "\u5b9f\u7e3e\u306e\u751f\u30c7\u30fc\u30bf\uff08\u4ef6\u6570\u30fb\u7387\uff09\u3092\u78ba\u8a8d\u3059\u308b\u30da\u30fc\u30b8\u3002\u70b9\u6570\u5316\u524d\u306e\u5b9f\u7e3e\u5024\u304c\u898b\u3089\u308c\u307e\u3059\u3002",
        "🏆 \u9806\u4f4d": "\u7dcf\u5408\u30b9\u30b3\u30a2\u306e\u9ad8\u3044\u9806\u306b\u30e9\u30f3\u30ad\u30f3\u30b0\u8868\u793a\u3002\u500b\u4eba\u30fb\u30c1\u30fc\u30e0\u30fb\u5e97\u8217\u306e\u30e9\u30f3\u30ad\u30f3\u30b0\u304c\u4e00\u89a7\u3067\u78ba\u8a8d\u3067\u304d\u307e\u3059\u3002",
        "📈 \u6210\u7e3e\u8a73\u7d30": "\u500b\u4eba\u30fb\u30c1\u30fc\u30e0\u30fb\u5e97\u8217\u3092\u9078\u3093\u3067\u3001\u9805\u76ee\u3054\u3068\u306e\u5185\u8a33\u30b9\u30b3\u30a2\u3068\u30ec\u30fc\u30c0\u30fc\u30c1\u30e3\u30fc\u30c8\u3092\u78ba\u8a8d\u3067\u304d\u307e\u3059\u3002",
        "📉 \u5206\u6790": "\u708e\u3001\u88c5\u5099\u306a\u3069\u306e\u30c7\u30fc\u30bf\u3092\u30b0\u30e9\u30d5\u3067\u53ef\u8996\u5316\u3057\u305f\u30d6\u30e9\u30f3\u30c9\u5206\u6790\u30da\u30fc\u30b8\u3002",
    }.items():
        st.markdown(f"**{page}** \u2014 {desc}")

# =====================================================
# 年間表彰ページ
# =====================================================
def annual_awards_page(df):
    st.header("年間表彰")
    st.markdown("各月の「満点に対する達成率」の平均で年間ランキングを決定します。月ごとの配点変動の影響を排除した**公平な評価**です。")

    import os
    from utils.data_manager import DATA_DIR, load_scoring_config as _load_cfg, load_performance_data

    # ── 年度選択 ──────────────────────────────────────
    st.sidebar.markdown("---")
    st.sidebar.markdown("**🏆 年間表彰 年度設定**")

    # 年度: 4月〜3月 を1年度とみなす
    # 例: 2025年度 = 2025年4月〜2026年3月
    from datetime import datetime
    now = datetime.now()
    # 当年度（4月始まり）
    current_fiscal_year = now.year if now.month >= 4 else now.year - 1

    # アーカイブフォルダから利用可能な年度を抽出
    available_dirs = []
    if os.path.exists(DATA_DIR):
        available_dirs = sorted([
            d for d in os.listdir(DATA_DIR)
            if os.path.isdir(os.path.join(DATA_DIR, d)) and d.isdigit() and len(d) == 6
        ])

    # アーカイブ年度リスト（4月始まり基準）
    fiscal_years = set()
    for d in available_dirs:
        y, m = int(d[:4]), int(d[4:])
        fy = y if m >= 4 else y - 1
        fiscal_years.add(fy)
    fiscal_years.add(current_fiscal_year)
    fiscal_year_list = sorted(fiscal_years, reverse=True)

    selected_fy = st.sidebar.selectbox(
        "表彰対象年度",
        fiscal_year_list,
        index=0,
        format_func=lambda y: f"{y}年度（{y}年4月〜{y+1}年3月）"
    )

    # 対象フォルダキーリスト（YYYYMMのリスト）
    def fy_months(fy):
        """年度の12ヶ月分のYYYYMMリスト（4月〜翌3月）"""
        months = []
        for m in range(4, 13):
            months.append(f"{fy}{str(m).zfill(2)}")
        for m in range(1, 4):
            months.append(f"{fy+1}{str(m).zfill(2)}")
        return months

    target_ym_list = fy_months(selected_fy)

    # ── 最低参加月数の設定 ───────────────────────────
    MIN_MONTHS = 6  # 表彰対象の最低参加月数

    st.markdown(f"### 📅 {selected_fy}年度（{selected_fy}年4月 〜 {selected_fy+1}年3月）")
    st.info(f"🏅 **表彰対象**: {MIN_MONTHS}ヶ月以上参加のメンバー　｜　未満のメンバーは参考値（🔘）として表示")

    # ── データ収集 ────────────────────────────────────
    # 各月のアーカイブから scored データを取得
    monthly_records = []  # [{name, shop, team, month_key, achievement_rate}]
    available_months = []

    status_placeholder = st.empty()

    for ym in target_ym_list:
        folder_path = os.path.join(DATA_DIR, ym)
        perf_path   = os.path.join(folder_path, 'performance_data.csv')
        cfg_path    = os.path.join(folder_path, 'scoring_config.json')

        if not os.path.exists(perf_path):
            continue

        try:
            month_df = pd.read_csv(perf_path)
            if month_df.empty:
                continue

            # その月の評価設定を使用（スナップショット）
            month_cfg = _load_cfg(month=ym) if os.path.exists(cfg_path) else load_scoring_config()
            if not month_cfg or not month_cfg.get('items'):
                continue

            # 満点の計算（Lv1 配点の合計）
            max_score = sum(item.get('weight', 0) for item in month_cfg.get('items', []))
            if max_score <= 0:
                continue

            df_denom = get_denominator_df(month_df)
            scored   = calculate_scores(month_df, month_cfg, df_denom)

            # スタッフ名列を特定
            n_col = 'スタッフ名' if 'スタッフ名' in scored.columns else 'Name'
            s_col = next((c for c in ['店舗名', '店舗'] if c in scored.columns), None)
            t_col = next((c for c in ['チーム名', 'チーム'] if c in scored.columns), None)

            if n_col not in scored.columns:
                continue

            for _, row in scored.iterrows():
                name  = row.get(n_col, '')
                if not name or str(name).strip() == '':
                    continue
                score = row.get('Total_Score', 0)
                rate  = round((score / max_score) * 100, 2)  # 達成率(%)

                monthly_records.append({
                    'name':      str(name).strip(),
                    'shop':      str(row.get(s_col, '')) if s_col else '',
                    'team':      str(row.get(t_col, '')) if t_col else '',
                    'month_key': ym,
                    'score':     score,
                    'max_score': max_score,
                    'rate':      rate,
                })

            available_months.append(ym)

        except Exception as e:
            st.warning(f"{ym} の読み込みエラー: {e}")
            continue

    status_placeholder.empty()

    if not monthly_records:
        st.warning(f"⚠️ {selected_fy}年度のバックナンバーデータが見つかりません。")
        st.markdown("管理者設定 → 成績データ管理 → **評価データの確定と保存** で各月のデータをアーカイブしてください。")
        return

    records_df = pd.DataFrame(monthly_records)

    # ── 集計関数 ─────────────────────────────────────
    def build_ranking(records, group_cols, label_col):
        """
        records: DataFrame(name, shop, team, month_key, rate)
        group_cols: グループ化する列のリスト
        label_col: 表示名として使う列
        """
        grp = records.groupby(group_cols + ['month_key'])['rate'].mean().reset_index()
        grp.columns = group_cols + ['month_key', 'rate']
        # 参加月数・平均達成率
        summary = grp.groupby(group_cols).agg(
            参加月数=('month_key', 'nunique'),
            年間平均達成率=('rate', 'mean')
        ).reset_index()
        summary['年間平均達成率'] = summary['年間平均達成率'].round(2)
        summary['表彰対象'] = summary['参加月数'] >= MIN_MONTHS

        # 月別達成率をピボット（参考列）
        pivot = grp.pivot_table(index=group_cols, columns='month_key', values='rate', aggfunc='mean')
        pivot.columns = [f"{c[4:6]}月" for c in pivot.columns]
        pivot = pivot.round(1).reset_index()

        merged = pd.merge(summary, pivot, on=group_cols, how='left')
        # 表彰対象のみで順位付け
        target = merged[merged['表彰対象']].sort_values('年間平均達成率', ascending=False).reset_index(drop=True)
        ref    = merged[~merged['表彰対象']].sort_values('年間平均達成率', ascending=False).reset_index(drop=True)

        target.insert(0, '順位', range(1, len(target)+1))
        ref.insert(0, '順位', ['—'] * len(ref))

        return target, ref, merged

    # ── タブ表示 ──────────────────────────────────────
    tab_ind, tab_team, tab_shop = st.tabs(["👤 個人", "🤝 チーム", "🏬 店舗"])

    def display_ranking_tab(target_df, ref_df, label_col, tab_label):
        """ランキングタブの共通表示"""

        total_months = len(available_months)
        st.caption(f"集計対象: {total_months}ヶ月 ／ 表彰条件: {MIN_MONTHS}ヶ月以上参加")

        # ── TOP3 表彰台 ──
        medals = ["🥇 1位", "🥈 2位", "🥉 3位"]
        top3 = target_df.head(3)
        if not top3.empty:
            cols = st.columns(len(top3))
            for i, (_, row) in enumerate(top3.iterrows()):
                with cols[i]:
                    st.markdown(f"### {medals[i]}")
                    st.markdown(f"**{row[label_col]}**")
                    st.metric("年間平均達成率", f"{row['年間平均達成率']:.2f}%")
                    st.caption(f"参加 {row['参加月数']}ヶ月")

        # ── 表彰対象テーブル ──
        st.markdown("#### 表彰対象ランキング")
        if not target_df.empty:
            st.dataframe(target_df, hide_index=True, use_container_width=True)
        else:
            st.info(f"{MIN_MONTHS}ヶ月以上参加のメンバーがいません。")

        # ── 参考値テーブル ──
        if not ref_df.empty:
            st.markdown(f"#### 🔘 参考値（{MIN_MONTHS}ヶ月未満）")
            st.caption("表彰対象外ですが達成率は参考として表示しています。")
            st.dataframe(ref_df, hide_index=True, use_container_width=True)

        # ── Excel DL ──
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            target_df.to_excel(writer, index=False, sheet_name='表彰対象')
            if not ref_df.empty:
                ref_df.to_excel(writer, index=False, sheet_name='参考値')
        st.download_button(
            label=f"📥 {tab_label}ランキングをExcelでダウンロード",
            data=buf.getvalue(),
            file_name=f"annual_{tab_label}_{selected_fy}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_{tab_label}"
        )

    with tab_ind:
        # ユーザーマスターでshop/teamを補完
        user_master = load_user_master()
        if user_master.get('users'):
            master_map = {u['name']: u for u in user_master['users']}
            records_df['shop'] = records_df.apply(
                lambda r: master_map.get(r['name'], {}).get('shop', r['shop']), axis=1
            )
            records_df['team'] = records_df.apply(
                lambda r: master_map.get(r['name'], {}).get('team', r['team']), axis=1
            )

        target_df, ref_df, _ = build_ranking(records_df, ['name', 'shop', 'team'], 'name')
        # 表示用に列名変換
        for d in [target_df, ref_df]:
            d.rename(columns={'name': 'スタッフ名', 'shop': '店舗', 'team': 'チーム'}, inplace=True)
        display_ranking_tab(target_df, ref_df, 'スタッフ名', '個人')

    with tab_team:
        team_records = records_df.copy()
        # チームごとの月別平均達成率を集計
        target_df, ref_df, _ = build_ranking(team_records, ['shop', 'team'], 'team')
        for d in [target_df, ref_df]:
            d.rename(columns={'shop': '店舗', 'team': 'チーム'}, inplace=True)
        display_ranking_tab(target_df, ref_df, 'チーム', 'チーム')

    with tab_shop:
        shop_records = records_df.copy()
        target_df, ref_df, _ = build_ranking(shop_records, ['shop'], 'shop')
        for d in [target_df, ref_df]:
            d.rename(columns={'shop': '店舗'}, inplace=True)
        display_ranking_tab(target_df, ref_df, '店舗', '店舗')



def _build_excel_bytes_for_download(dl_df, dl_config, dl_target_month):
    """個人・チーム・店舗テーブルを1シートに縦積みしてバイト列を返す（トップレベル関数）"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    if dl_df is None or dl_df.empty or not dl_config or not dl_config.get('items'):
        return None, "データまたは設定が不足しています"
    try:
        from utils.calculation import normalize_col_name as _ncn_dl
        dl_denom = get_denominator_df(dl_df)
        scored = calculate_scores(dl_df, dl_config, dl_denom)
        items_dl = dl_config.get('items', [])
        item_names_dl = [it.get('name', '') for it in items_dl if it.get('name')]

        name_col_dl = 'スタッフ名' if 'スタッフ名' in scored.columns else 'Name'
        shop_col_dl = next((c for c in ['店舗名', '店舗'] if c in scored.columns), None)
        team_col_dl = next((c for c in ['チーム名', 'チーム'] if c in scored.columns), None)

        score_cols = [f"{_ncn_dl(n)}_score" for n in item_names_dl]
        score_cols_exist = [c for c in score_cols if c in scored.columns]
        rename_map = {f"{_ncn_dl(n)}_score": n for n in item_names_dl}

        # --- 個人テーブル（名簿順） ---
        ind_base_cols = [c for c in [name_col_dl, shop_col_dl, team_col_dl] if c]
        ind_df = scored[ind_base_cols + ['Rank', 'Total_Score'] + score_cols_exist].copy()
        ind_df.rename(columns={'Rank': '順位', 'Total_Score': '合計スコア'}, inplace=True)
        ind_df.rename(columns=rename_map, inplace=True)
        um_dl = load_user_master(month=dl_target_month)
        ind_df = apply_robust_sorting(ind_df, um_dl, name_col_dl)
        for c in ind_df.select_dtypes('float').columns:
            ind_df[c] = ind_df[c].round(1)

        # --- チームテーブル ---
        team_scored = aggregate_team_scores(scored, dl_df, dl_config)
        team_base = [c for c in ['チーム名', '店舗名'] if c in team_scored.columns]
        team_score_cols_e = [c for c in score_cols if c in team_scored.columns]
        team_df = team_scored[team_base + ['Rank', 'Total_Score'] + team_score_cols_e].copy()
        team_df.rename(columns={'Rank': '順位', 'Total_Score': '合計スコア'}, inplace=True)
        team_df.rename(columns=rename_map, inplace=True)
        team_df = team_df.sort_values('順位').reset_index(drop=True)
        for c in team_df.select_dtypes('float').columns:
            team_df[c] = team_df[c].round(1)

        # --- 店舗テーブル ---
        shop_scored = aggregate_shop_scores(scored, dl_df, dl_config)
        shop_base = [c for c in ['店舗名'] if c in shop_scored.columns]
        shop_score_cols_e = [c for c in score_cols if c in shop_scored.columns]
        shop_df = shop_scored[shop_base + ['Rank', 'Total_Score'] + shop_score_cols_e].copy()
        shop_df.rename(columns={'Rank': '順位', 'Total_Score': '合計スコア'}, inplace=True)
        shop_df.rename(columns=rename_map, inplace=True)
        shop_df = shop_df.sort_values('順位').reset_index(drop=True)
        for c in shop_df.select_dtypes('float').columns:
            shop_df[c] = shop_df[c].round(1)

        # --- openpyxlで1シートに縦積み書き込み ---
        buf_dl = io.BytesIO()
        wb_dl = openpyxl.Workbook()
        ws_dl = wb_dl.active
        ws_dl.title = '成績一覧'

        def _write_dl_section(ws, label, df):
            label_row = [label] + [''] * (len(df.columns) - 1)
            ws.append(label_row)
            ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
            ws.append(list(df.columns))
            for c_idx in range(1, len(df.columns) + 1):
                ws.cell(ws.max_row, c_idx).font = Font(bold=True)
            for _, row_data in df.iterrows():
                ws.append([
                    None if (isinstance(v, float) and pd.isna(v)) else v
                    for v in row_data.tolist()
                ])
            ws.append([])
            ws.append([])

        _write_dl_section(ws_dl, '【個人】', ind_df)
        _write_dl_section(ws_dl, '【チーム】', team_df)
        _write_dl_section(ws_dl, '【店舗】', shop_df)

        wb_dl.save(buf_dl)
        return buf_dl.getvalue(), None

    except Exception as e:
        import traceback
        return None, f"{e}\n{traceback.format_exc()}"


def main():
    st.set_page_config(page_title="ブライトパスくん", layout="wide")

    # ============================================================
    # ログイン認証ゲート（ページ全体を保護）
    # ============================================================

    # ログインフォームを画面中央に配置するCSS / ログアウトボタンの中央揃えCSS
    st.markdown("""
    <style>
        /* ログインフォームを中央配置・幅を制限 */
        [data-testid="stForm"] {
            max-width: 400px !important;
            margin: 0 auto !important;
        }
        /* ログアウトボタン: サイドバー内の全ボタンを幅いっぱいに */
        section[data-testid="stSidebar"] button {
            width: 100% !important;
            text-align: center !important;
            display: flex !important;
            justify-content: center !important;
            align-items: center !important;
        }
    </style>
    """, unsafe_allow_html=True)

    _auth_config = load_auth_config()
    authenticator = stauth.Authenticate(
        _auth_config["credentials"],
        _auth_config["cookie"]["name"],
        _auth_config["cookie"]["key"],
        _auth_config["cookie"]["expiry_days"],
    )

    # ログイン画面を表示
    name, authentication_status, username = authenticator.login(
        location="main",
        fields={
            "Form name": "ブライトパスくん ログイン",
            "Username": "ユーザーID",
            "Password": "パスワード",
            "Login": "ログイン",
        }
    )

    if authentication_status is False:
        # ログイン失敗
        log_login(username or "(不明)", "", "failed", "パスワードまたはIDが違います")
        st.error("❌ ユーザーIDまたはパスワードが違います")
        st.stop()
    elif authentication_status is None:
        # 未入力（初期表示）
        st.warning("ユーザーIDとパスワードを入力してください")
        st.stop()
    else:
        # ログイン成功
        user_role = _auth_config["credentials"]["usernames"].get(username, {}).get("role", "viewer")
        # セッションにロールを保存
        if st.session_state.get("_logged_username") != username:
            # 新規ログイン時のみログ記録（ページリロードで二重記録しない）
            log_login(username, name, "success", "ログイン成功")
            st.session_state["_logged_username"] = username
        st.session_state["user_role"] = user_role
        st.session_state["auth_name"] = name
        st.session_state["auth_username"] = username

        # ログアウトボタン（サイドバー）
        # authenticator.logout() を使うことでクッキーも正しく削除される
        authenticator.logout("ログアウト", location="sidebar", key="logout_btn")

    st.title("ブライトパスくん")

    # CSS injection
    # CSS injection for Material Design & Strict Layout
    st.markdown("""
    <style>
        /* =============================================
           ブライトパスくん デザインシステム
           オーガニック・マテリアルデザイン
           ============================================= */

        /* Google Fonts */
        @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+JP:wght@300;400;500;700&display=swap');

        /* ベーススタイル */
        html, body {
            font-family: 'Noto Sans JP', sans-serif !important;
        }

        /* メインエリア背景 */
        .main, [data-testid="stAppViewContainer"] > .main {
            background-color: #F5F8FA !important;
        }
        [data-testid="stAppViewContainer"] {
            background-color: #F5F8FA !important;
        }

        /* 見出し */
        h1 { color: #2C3E50 !important; font-weight: 700 !important; font-family: 'Noto Sans JP', sans-serif !important; letter-spacing: 0.03em !important; }
        h2, h3, h4 { color: #2C3E50 !important; font-weight: 600 !important; font-family: 'Noto Sans JP', sans-serif !important; }

        /* =====================
           ボタン（プライマリ）
           ===================== */
        .stButton > button[kind="primary"],
        .stButton > button[data-testid="baseButton-primary"] {
            background-color: #0097B2 !important;
            color: #FFFFFF !important;
            border: none !important;
            border-radius: 20px !important;
            padding: 10px 22px !important;
            font-weight: 600 !important;
            font-family: 'Noto Sans JP', sans-serif !important;
            font-size: 0.88rem !important;
            letter-spacing: 0.04em !important;
            box-shadow: 0 2px 8px rgba(0,151,178,0.22) !important;
            transition: all 0.2s ease !important;
        }
        .stButton > button[kind="primary"]:hover,
        .stButton > button[data-testid="baseButton-primary"]:hover {
            background-color: #007A91 !important;
            box-shadow: 0 4px 16px rgba(0,151,178,0.35) !important;
            transform: translateY(-1px) !important;
        }

        /* ボタン（セカンダリ） */
        .stButton > button[kind="secondary"],
        .stButton > button[data-testid="baseButton-secondary"] {
            background-color: #FFFFFF !important;
            color: #0097B2 !important;
            border: 1.5px solid #A8D8E3 !important;
            border-radius: 20px !important;
            padding: 10px 22px !important;
            font-weight: 500 !important;
            font-family: 'Noto Sans JP', sans-serif !important;
            font-size: 0.88rem !important;
            letter-spacing: 0.03em !important;
            transition: all 0.2s ease !important;
        }
        .stButton > button[kind="secondary"]:hover,
        .stButton > button[data-testid="baseButton-secondary"]:hover {
            background-color: #EDF7FA !important;
            border-color: #0097B2 !important;
            color: #007A91 !important;
        }

        /* =====================
           サイドバー
           ===================== */
        section[data-testid="stSidebar"] {
            background-color: #EBF3F6 !important;
        }
        section[data-testid="stSidebar"] > div:first-child {
            background-color: #EBF3F6 !important;
            padding-top: 1.5rem !important;
        }

        /* サイドバー ボタン共通 */
        section[data-testid="stSidebar"] .stButton > button {
            border-radius: 16px !important;
            padding: 10px 14px !important;
            width: 100% !important;
            margin-bottom: 4px !important;
            font-size: 0.88rem !important;
            font-weight: 500 !important;
            letter-spacing: 0.03em !important;
            text-align: center !important;
            display: flex !important;
            justify-content: center !important;
            align-items: center !important;
            transition: all 0.18s ease !important;
        }

        /* サイドバー ボタン内部の全要素を全幅・中央揃え */
        section[data-testid="stSidebar"] .stButton > button > div,
        section[data-testid="stSidebar"] .stButton > button > div > span,
        section[data-testid="stSidebar"] .stButton > button > div > span > div {
            width: 100% !important;
            display: flex !important;
            justify-content: center !important;
            align-items: center !important;
        }

        /* サイドバー ボタン p要素（テキスト中央揃え） */
        section[data-testid="stSidebar"] .stButton > button p {
            width: 100% !important;
            text-align: center !important;
            display: block !important;
            margin: 0 !important;
            padding: 0 !important;
        }

        /* SVGアイコン（::before）を非表示 - テキスト中央揃えのため */
        section[data-testid="stSidebar"] .stButton > button p::before {
            display: none !important;
            content: none !important;
        }

        /* 総合 - バーチャートアイコン */
        section[data-testid="stSidebar"] button[aria-label="総合"] p::before {
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%234A4A4A' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Crect x='3' y='12' width='4' height='9'/%3E%3Crect x='10' y='6' width='4' height='15'/%3E%3Crect x='17' y='3' width='4' height='18'/%3E%3C/svg%3E") !important;
        }
        section[data-testid="stSidebar"] button[kind="primary"][aria-label="総合"] p::before {
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23FFFFFF' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Crect x='3' y='12' width='4' height='9'/%3E%3Crect x='10' y='6' width='4' height='15'/%3E%3Crect x='17' y='3' width='4' height='18'/%3E%3C/svg%3E") !important;
        }

        /* 件数 - リストアイコン */
        section[data-testid="stSidebar"] button[aria-label="件数"] p::before {
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%234A4A4A' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Cline x1='8' y1='6' x2='21' y2='6'/%3E%3Cline x1='8' y1='12' x2='21' y2='12'/%3E%3Cline x1='8' y1='18' x2='21' y2='18'/%3E%3Ccircle cx='3' cy='6' r='1.5' fill='%234A4A4A'/%3E%3Ccircle cx='3' cy='12' r='1.5' fill='%234A4A4A'/%3E%3Ccircle cx='3' cy='18' r='1.5' fill='%234A4A4A'/%3E%3C/svg%3E") !important;
        }
        section[data-testid="stSidebar"] button[kind="primary"][aria-label="件数"] p::before {
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23FFFFFF' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Cline x1='8' y1='6' x2='21' y2='6'/%3E%3Cline x1='8' y1='12' x2='21' y2='12'/%3E%3Cline x1='8' y1='18' x2='21' y2='18'/%3E%3Ccircle cx='3' cy='6' r='1.5' fill='%23FFFFFF'/%3E%3Ccircle cx='3' cy='12' r='1.5' fill='%23FFFFFF'/%3E%3Ccircle cx='3' cy='18' r='1.5' fill='%23FFFFFF'/%3E%3C/svg%3E") !important;
        }

        /* 順位 - トロフィーアイコン */
        section[data-testid="stSidebar"] button[aria-label="順位"] p::before {
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%234A4A4A' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpath d='M6 9H3a1 1 0 0 1-1-1V5a1 1 0 0 1 1-1h3'/%3E%3Cpath d='M18 9h3a1 1 0 0 0 1-1V5a1 1 0 0 0-1-1h-3'/%3E%3Cpath d='M6 4h12v8a6 6 0 0 1-12 0V4z'/%3E%3Cpath d='M12 18v4'/%3E%3Cpath d='M8 22h8'/%3E%3C/svg%3E") !important;
        }
        section[data-testid="stSidebar"] button[kind="primary"][aria-label="順位"] p::before {
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23FFFFFF' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpath d='M6 9H3a1 1 0 0 1-1-1V5a1 1 0 0 1 1-1h3'/%3E%3Cpath d='M18 9h3a1 1 0 0 0 1-1V5a1 1 0 0 0-1-1h-3'/%3E%3Cpath d='M6 4h12v8a6 6 0 0 1-12 0V4z'/%3E%3Cpath d='M12 18v4'/%3E%3Cpath d='M8 22h8'/%3E%3C/svg%3E") !important;
        }

        /* 成績詳細 - 人物アイコン */
        section[data-testid="stSidebar"] button[aria-label="成績詳細"] p::before {
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%234A4A4A' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Ccircle cx='12' cy='8' r='4'/%3E%3Cpath d='M20 21a8 8 0 1 0-16 0'/%3E%3C/svg%3E") !important;
        }
        section[data-testid="stSidebar"] button[kind="primary"][aria-label="成績詳細"] p::before {
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23FFFFFF' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Ccircle cx='12' cy='8' r='4'/%3E%3Cpath d='M20 21a8 8 0 1 0-16 0'/%3E%3C/svg%3E") !important;
        }

        /* 分析 - 折れ線グラフアイコン */
        section[data-testid="stSidebar"] button[aria-label="分析"] p::before {
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%234A4A4A' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpolyline points='22 12 18 12 15 21 9 3 6 12 2 12'/%3E%3C/svg%3E") !important;
        }
        section[data-testid="stSidebar"] button[kind="primary"][aria-label="分析"] p::before {
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23FFFFFF' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpolyline points='22 12 18 12 15 21 9 3 6 12 2 12'/%3E%3C/svg%3E") !important;
        }

        /* ルール説明 - ドキュメントアイコン */
        section[data-testid="stSidebar"] button[aria-label="ルール説明"] p::before {
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%234A4A4A' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpath d='M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z'/%3E%3Cpolyline points='14 2 14 8 20 8'/%3E%3Cline x1='16' y1='13' x2='8' y2='13'/%3E%3Cline x1='16' y1='17' x2='8' y2='17'/%3E%3C/svg%3E") !important;
        }
        section[data-testid="stSidebar"] button[kind="primary"][aria-label="ルール説明"] p::before {
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23FFFFFF' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpath d='M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z'/%3E%3Cpolyline points='14 2 14 8 20 8'/%3E%3Cline x1='16' y1='13' x2='8' y2='13'/%3E%3Cline x1='16' y1='17' x2='8' y2='17'/%3E%3C/svg%3E") !important;
        }

        /* 配点シミュレーション - スライダーアイコン */
        section[data-testid="stSidebar"] button[aria-label="配点シミュレーション"] p::before {
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%234A4A4A' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Cline x1='4' y1='6' x2='20' y2='6'/%3E%3Cline x1='4' y1='12' x2='20' y2='12'/%3E%3Cline x1='4' y1='18' x2='20' y2='18'/%3E%3Ccircle cx='8' cy='6' r='2' fill='%23F5F8FA'/%3E%3Ccircle cx='15' cy='12' r='2' fill='%23F5F8FA'/%3E%3Ccircle cx='10' cy='18' r='2' fill='%23F5F8FA'/%3E%3C/svg%3E") !important;
        }
        section[data-testid="stSidebar"] button[kind="primary"][aria-label="配点シミュレーション"] p::before {
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23FFFFFF' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Cline x1='4' y1='6' x2='20' y2='6'/%3E%3Cline x1='4' y1='12' x2='20' y2='12'/%3E%3Cline x1='4' y1='18' x2='20' y2='18'/%3E%3Ccircle cx='8' cy='6' r='2' fill='%230097B2'/%3E%3Ccircle cx='15' cy='12' r='2' fill='%230097B2'/%3E%3Ccircle cx='10' cy='18' r='2' fill='%230097B2'/%3E%3C/svg%3E") !important;
        }

        /* 年間表彰 - スターアイコン */
        section[data-testid="stSidebar"] button[aria-label="年間表彰"] p::before {
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%234A4A4A' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpolygon points='12 2 15.09 8.26 22 9.27 17 14.14 18.18 21.02 12 17.77 5.82 21.02 7 14.14 2 9.27 8.91 8.26 12 2'/%3E%3C/svg%3E") !important;
        }
        section[data-testid="stSidebar"] button[kind="primary"][aria-label="年間表彰"] p::before {
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23FFFFFF' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpolygon points='12 2 15.09 8.26 22 9.27 17 14.14 18.18 21.02 12 17.77 5.82 21.02 7 14.14 2 9.27 8.91 8.26 12 2'/%3E%3C/svg%3E") !important;
        }

        /* 設定 - 歯車アイコン */
        section[data-testid="stSidebar"] button[aria-label="設定"] p::before {
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%234A4A4A' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Ccircle cx='12' cy='12' r='3'/%3E%3Cpath d='M19.4 15a1.65 1.65 0 0 0 .33 1.82l.06.06a2 2 0 0 1-2.83 2.83l-.06-.06a1.65 1.65 0 0 0-1.82-.33 1.65 1.65 0 0 0-1 1.51V21a2 2 0 0 1-4 0v-.09A1.65 1.65 0 0 0 9 19.4a1.65 1.65 0 0 0-1.82.33l-.06.06a2 2 0 0 1-2.83-2.83l.06-.06A1.65 1.65 0 0 0 4.68 15a1.65 1.65 0 0 0-1.51-1H3a2 2 0 0 1 0-4h.09A1.65 1.65 0 0 0 4.6 9a1.65 1.65 0 0 0-.33-1.82l-.06-.06a2 2 0 0 1 2.83-2.83l.06.06A1.65 1.65 0 0 0 9 4.68a1.65 1.65 0 0 0 1-1.51V3a2 2 0 0 1 4 0v.09a1.65 1.65 0 0 0 1 1.51 1.65 1.65 0 0 0 1.82-.33l.06-.06a2 2 0 0 1 2.83 2.83l-.06.06A1.65 1.65 0 0 0 19.4 9a1.65 1.65 0 0 0 1.51 1H21a2 2 0 0 1 0 4h-.09a1.65 1.65 0 0 0-1.51 1z'/%3E%3C/svg%3E") !important;
        }
        section[data-testid="stSidebar"] button[kind="primary"][aria-label="設定"] p::before {
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='%23FFFFFF' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Ccircle cx='12' cy='12' r='3'/%3E%3Cpath d='M19.4 15a1.65 1.65 0 0 0 .33 1.82l.06.06a2 2 0 0 1-2.83 2.83l-.06-.06a1.65 1.65 0 0 0-1.82-.33 1.65 1.65 0 0 0-1 1.51V21a2 2 0 0 1-4 0v-.09A1.65 1.65 0 0 0 9 19.4a1.65 1.65 0 0 0-1.82.33l-.06.06a2 2 0 0 1-2.83-2.83l.06-.06A1.65 1.65 0 0 0 4.68 15a1.65 1.65 0 0 0-1.51-1H3a2 2 0 0 1 0-4h.09A1.65 1.65 0 0 0 4.6 9a1.65 1.65 0 0 0-.33-1.82l-.06-.06a2 2 0 0 1 2.83-2.83l.06.06A1.65 1.65 0 0 0 9 4.68a1.65 1.65 0 0 0 1-1.51V3a2 2 0 0 1 4 0v.09a1.65 1.65 0 0 0 1 1.51 1.65 1.65 0 0 0 1.82-.33l.06-.06a2 2 0 0 1 2.83 2.83l-.06.06A1.65 1.65 0 0 0 19.4 9a1.65 1.65 0 0 0 1.51 1H21a2 2 0 0 1 0 4h-.09a1.65 1.65 0 0 0-1.51 1z'/%3E%3C/svg%3E") !important;
        }

        /* =====================
           タブ
           ===================== */
        [data-testid="stTabs"] {
            background-color: transparent !important;
        }
        [data-testid="stTabs"] button {
            border-radius: 20px !important;
            border: 1.5px solid #A8D8E3 !important;
            background-color: #FFFFFF !important;
            color: #4A4A4A !important;
            padding: 8px 20px !important;
            margin-right: 8px !important;
            margin-bottom: 12px !important;
            font-size: 0.88rem !important;
            font-weight: 500 !important;
            font-family: 'Noto Sans JP', sans-serif !important;
            transition: all 0.2s ease !important;
            height: auto !important;
            min-height: 40px !important;
        }
        [data-testid="stTabs"] button:hover {
            background-color: #EDF7FA !important;
            border-color: #0097B2 !important;
            color: #0097B2 !important;
        }
        [data-testid="stTabs"] button[aria-selected="true"] {
            background-color: #0097B2 !important;
            color: #FFFFFF !important;
            border-color: #0097B2 !important;
            box-shadow: 0 2px 8px rgba(0,151,178,0.25) !important;
        }
        [data-testid="stTabs"] button[aria-selected="true"] p {
            color: #FFFFFF !important;
        }
        div[data-baseweb="tab-highlight"] { display: none !important; }

        /* =====================
           テーブル / データフレーム
           ===================== */
        [data-testid="stDataFrame"] {
            background-color: #FFFFFF;
            border-radius: 12px !important;
            box-shadow: 0 2px 12px rgba(0,0,0,0.06) !important;
            border: none !important;
            overflow: hidden;
            width: 100% !important;
        }
        [data-testid="stDataFrame"] > div > div {
            border-radius: 12px !important;
        }
        thead tr th {
            color: #7F8C8D !important;
            font-size: 0.82rem !important;
            font-weight: 600 !important;
            letter-spacing: 0.04em !important;
            background-color: #F5F8FA !important;
            border-bottom: 1.5px solid #E0E8EB !important;
        }

        /* =====================
           ログインフォーム
           ===================== */
        [data-testid="stForm"] {
            max-width: 420px !important;
            margin: 2rem auto !important;
            background-color: #FFFFFF !important;
            border-radius: 16px !important;
            padding: 2rem !important;
            box-shadow: 0 4px 24px rgba(0,0,0,0.08) !important;
            border: none !important;
        }

        /* =====================
           ダウンロードボタン
           ===================== */
        .stDownloadButton > button {
            border-radius: 20px !important;
            font-weight: 600 !important;
            background-color: #FFFFFF !important;
            color: #0097B2 !important;
            border: 1.5px solid #A8D8E3 !important;
            width: 100% !important;
            transition: all 0.2s ease !important;
        }
        .stDownloadButton > button:hover {
            background-color: #EDF7FA !important;
            border-color: #0097B2 !important;
        }

        /* =====================
           ファイルアップローダー
           ===================== */
        [data-testid="stFileUploaderDropzone"] {
            background-color: #FFFFFF;
            border: 1.5px dashed #0097B2;
            border-radius: 12px;
        }
        [data-testid="stFileUploaderDropzone"] button {
            border-radius: 16px !important;
            background-color: #FFFFFF !important;
            color: #0097B2 !important;
            box-shadow: none !important;
            border: 1.5px solid #A8D8E3 !important;
        }
        /* ファイルアップローダー日本語化 */
        [data-testid="stFileUploadDropzone"] > div > small,
        [data-testid="stFileUploaderDropzone"] > div > small,
        [data-testid="stFileUploadDropzone"] > div > div > span,
        [data-testid="stFileUploaderDropzone"] > div > div > span {
            display: none !important;
        }
        [data-testid="stFileUploadDropzone"] > div::before,
        [data-testid="stFileUploaderDropzone"] > div::before {
            content: "エクセルをここにドラッグ＆ドロップ (または右のボタンから選択)";
            display: block;
            text-align: center;
            font-size: 14px;
            color: #7F8C8D;
            margin-bottom: 8px;
        }

        /* =====================
           アラート
           ===================== */
        [data-testid="stAlert"] {
            border-radius: 12px !important;
            border: none !important;
        }

        /* Sticky Table Header */
        .report-container {
            max-height: 70vh !important;
        }
    </style>
    """, unsafe_allow_html=True)
    # --- Configuration loading ---
    config = load_scoring_config()
    user_master = load_user_master()

    # --- Sidebar: Back Number Selection ---
    st.sidebar.header("バックナンバー選択")

    from datetime import datetime
    now = datetime.now()
    current_year  = now.year
    current_month = now.month

    # data/ 配下の YYYYMM フォルダをスキャン
    archives = []
    if os.path.exists(DATA_DIR):
        archives = sorted(
            [d for d in os.listdir(DATA_DIR)
             if os.path.isdir(os.path.join(DATA_DIR, d)) and d.isdigit() and len(d) == 6],
            reverse=True
        )

    # 年度リスト作成（アーカイブの年 ＋ 当年）
    archive_years = sorted(set(int(d[:4]) for d in archives), reverse=True)
    if current_year not in archive_years:
        archive_years = [current_year] + archive_years

    # 年度セレクト（デフォルト: 当年）
    default_year_idx = archive_years.index(current_year) if current_year in archive_years else 0
    selected_year = st.sidebar.selectbox(
        "年度",
        archive_years,
        index=default_year_idx,
        format_func=lambda y: f"{y}年"
    )

    # 月リスト作成: 選択年度のアーカイブ月 ＋ 当年当月を「最新」として追加
    year_archives = [d for d in archives if d.startswith(str(selected_year))]
    # 月番号リスト (1〜12)
    archive_months_in_year = sorted(set(int(d[4:]) for d in year_archives), reverse=True)
    # 当年選択時は当月を「最新（現在の設定）」として先頭に追加
    month_options = []  # (表示ラベル, フォルダキー or None)
    if selected_year == current_year:
        month_options.append((f"{current_month}月（最新・現在の設定）", None))
    for m in archive_months_in_year:
        # 当月アーカイブは当年の場合スキップ（最新で代替）
        if selected_year == current_year and m == current_month:
            continue
        folder_key = f"{selected_year}{str(m).zfill(2)}"
        month_options.append((f"{m}月", folder_key))
    # 当年以外で月の選択肢がない場合のフォールバック
    if not month_options:
        month_options.append(("（データなし）", None))

    # デフォルト: 当年なら当月（最新）= index 0、それ以外は最初のアーカイブ
    selected_month_label, selected_folder = st.sidebar.selectbox(
        "月",
        month_options,
        index=0,
        format_func=lambda x: x[0]
    )

    target_month = selected_folder  # None なら最新（現在の設定）

    # 選択された年月に基づいて設定とデータを再ロード
    if target_month:
        config      = load_scoring_config(month=target_month)
        user_master = load_user_master(month=target_month)
        archived_df = load_performance_data(month=target_month)
        if archived_df is not None:
            st.session_state.performance_df = archived_df
            st.sidebar.success(f"✅ {target_month[:4]}年{int(target_month[4:])}月のデータを表示中")
        else:
            st.sidebar.warning("⚠️ 実績データが見つかりません。")


    # --- Sidebar ---
    st.sidebar.header("メニュー")

    # ログインユーザー情報をサイドバーに表示
    _auth_name_disp = st.session_state.get("auth_name", "")
    _auth_role_disp = st.session_state.get("user_role", "viewer")
    # ロールラベルを3段階で切り替え
    if _auth_role_disp == "superadmin":
        _role_label = "🔑 最高管理者"
    elif _auth_role_disp == "admin":
        _role_label = "👑 管理者"
    else:
        _role_label = "👤 スタッフ"
    st.sidebar.markdown(
        f"<div style='font-size:0.85rem; color:#5F6368; margin-bottom:4px;'>ログイン中: <b>{_auth_name_disp}</b>　{_role_label}</div>",
        unsafe_allow_html=True
    )
    st.sidebar.markdown("---")

    # ロールベースで is_admin を決定（旧パスワード認証の代替）
    # admin または superadmin の場合に管理者メニューを表示
    st.session_state['is_admin'] = (_auth_role_disp in ("admin", "superadmin"))

    # State for Mode
    if 'current_mode' not in st.session_state:
        st.session_state['current_mode'] = "総合"

    st.sidebar.markdown("<p style='color: #5F6368; font-weight: bold; font-size: 0.9rem; margin-bottom: 0px;'>モード選択</p>", unsafe_allow_html=True)
    
    # Custom Button Navigation
    if st.sidebar.button("総合", use_container_width=True, type="primary" if st.session_state['current_mode'] == "総合" else "secondary"):
        st.session_state['current_mode'] = "総合"
        st.rerun()
        
    if st.sidebar.button("件数", use_container_width=True, type="primary" if st.session_state['current_mode'] == "件数" else "secondary"):
        st.session_state['current_mode'] = "件数"
        st.rerun()
        
    if st.sidebar.button("順位", use_container_width=True, type="primary" if st.session_state['current_mode'] == "順位" else "secondary"):
        st.session_state['current_mode'] = "順位"
        st.rerun()

    if st.sidebar.button("成績詳細", use_container_width=True, type="primary" if st.session_state['current_mode'] == "成績詳細" else "secondary"):
        st.session_state['current_mode'] = "成績詳細"
        st.rerun()

    if st.sidebar.button("分析", use_container_width=True, type="primary" if st.session_state['current_mode'] == "分析" else "secondary"):
        st.session_state['current_mode'] = "分析"
        st.rerun()

    if st.sidebar.button("ルール説明", use_container_width=True, type="primary" if st.session_state['current_mode'] == "ルール説明" else "secondary"):
        st.session_state['current_mode'] = "ルール説明"
        st.rerun()

    if st.session_state['is_admin']:
        if st.sidebar.button("配点シミュレーション", use_container_width=True, type="primary" if st.session_state['current_mode'] == "配点シミュレーション" else "secondary"):
            st.session_state['current_mode'] = "配点シミュレーション"
            st.rerun()
        if st.sidebar.button("年間表彰", use_container_width=True, type="primary" if st.session_state['current_mode'] == "年間表彰" else "secondary"):
            st.session_state['current_mode'] = "年間表彰"
            st.rerun()
        
    if st.session_state['is_admin']:
        if st.sidebar.button("設定", use_container_width=True, type="primary" if st.session_state['current_mode'] == "設定 (Admin)" else "secondary"):
            st.session_state['current_mode'] = "設定 (Admin)"
            st.rerun()

    # CX分配くんボタン（最高管理者専用）
    if st.session_state.get("user_role") == "superadmin":
        if st.sidebar.button("💰 CX分配くん", use_container_width=True, type="primary" if st.session_state['current_mode'] == "CX分配くん" else "secondary"):
            st.session_state['current_mode'] = "CX分配くん"
            st.rerun()

    # --- サイドバー最下部: 一括Excelダウンロード ---
    st.sidebar.markdown("---")

    # データが存在する場合のみダウンロードボタンを表示
    _dl_perf_path = (
        os.path.join(DATA_DIR, selected_folder, "performance_data.csv")
        if 'selected_folder' in dir() and selected_folder
        else os.path.join(DATA_DIR, "current_performance.csv")
    )
    if os.path.exists(_dl_perf_path):
        try:
            _dl_df = pd.read_csv(_dl_perf_path)
            _dl_config = load_scoring_config(month=selected_folder if selected_folder else None)
            _ym_label = (
                f"{target_month[:4]}年{int(target_month[4:])}月"
                if target_month else "最新"
            )
            _fname = (
                f"ブライトパスくん_成績一覧_{target_month}.xlsx"
                if target_month else f"ブライトパスくん_成績一覧_最新.xlsx"
            )
            _excel_bytes, _dl_err = _build_excel_bytes_for_download(_dl_df, _dl_config, target_month)
            if _dl_err:
                st.sidebar.error(f'Excelエラー: {_dl_err[:300]}')
            elif _excel_bytes:
                st.sidebar.download_button(
                    label=f"📥 {_ym_label}データを一括DL",
                    data=_excel_bytes,
                    file_name=_fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="sidebar_bulk_download"
                )
        except Exception as _ex:
            st.sidebar.error(f"DLエラー: {_ex}")

    # --- サイドバー: 成績データ最終更新日時を表示 ---
    if os.path.exists(_dl_perf_path):
        try:
            from datetime import datetime as _dt
            _mtime = os.path.getmtime(_dl_perf_path)
            _mtime_str = _dt.fromtimestamp(_mtime).strftime("%Y/%m/%d %H:%M:%S")
            st.sidebar.markdown(
                f"<div style='font-size:0.8rem; color:#5F6368; margin-top:8px;'>"
                f"📊 データ最終更新<br><b>{_mtime_str}</b></div>",
                unsafe_allow_html=True
            )
        except Exception:
            pass

    mode = st.session_state['current_mode']

    # --- Global Data Loading ---
    # バックナンバー選択時はアーカイブのCSVを使用、最新月は current_performance.csv を使用
    df = None
    if target_month:
        # 過去月選択: data/YYYYMM/performance_data.csv を読み込む
        archived_path = os.path.join(DATA_DIR, target_month, "performance_data.csv")
        if os.path.exists(archived_path):
            try:
                df = pd.read_csv(archived_path, encoding='utf-8-sig')
                if 'スタッフ名' in df.columns:
                    df['スタッフ名'] = df['スタッフ名'].astype(str)
            except Exception as e:
                st.sidebar.error(f"アーカイブデータ読込エラー: {e}")
        else:
            if mode != "設定 (Admin)":
                st.sidebar.warning("⚠️ このアーカイブに実績データが見つかりません。")
    else:
        # 最新月: data/current_performance.csv を読み込む
        data_path = os.path.join(DATA_DIR, "current_performance.csv")
        if os.path.exists(data_path):
            try:
                df = pd.read_csv(data_path)
                if 'スタッフ名' in df.columns:
                    df['スタッフ名'] = df['スタッフ名'].astype(str)
            except Exception as e:
                st.sidebar.error(f"データ読込エラー: {e}")
        else:
            if mode != "設定 (Admin)":
                st.sidebar.warning("成績データが未登録です。管理者メニューからアップロードしてください。")


    # --- ページ上部: 成績データ最終更新日時を表示（閲覧ページのみ） ---
    _view_modes = {"総合", "順位", "件数", "成績詳細", "分析", "年間表彰"}
    if mode in _view_modes and df is not None and os.path.exists(_dl_perf_path):
        try:
            from datetime import datetime as _dt2
            _mtime2 = os.path.getmtime(_dl_perf_path)
            _mtime2_str = _dt2.fromtimestamp(_mtime2).strftime("%Y/%m/%d %H:%M:%S")
            st.caption(f"📊 データ最終更新: {_mtime2_str}")
        except Exception:
            pass

    # --- Routing ---
    if mode == "設定 (Admin)":
        admin_page(config, user_master, target_month=target_month)
    elif mode == "順位":
        ranking_page(df, target_month=target_month)
    elif mode == "件数":
        count_page(df, target_month=target_month)
    elif mode == "成績詳細":
        individual_detail_page(df, target_month=target_month)
    elif mode == "分析":
        analysis_page(df)
    elif mode == "配点シミュレーション":
        simulation_page(df)
    elif mode == "年間表彰":
        annual_awards_page(df)
    elif mode == "ルール説明":
        rules_page(config)
    elif mode == "CX分配くん":
        # 既存ロジックの再利用（読み込みのみ）
        cx_distributor_page(df, config, user_master, target_month=target_month)
    else:
        comprehensive_page(df, target_month=target_month)

# ============================================================
# CX分配くん ページ（管理者専用 / 既存ロジック不変）
# ============================================================
def cx_distributor_page(df, config, user_master, target_month=None):
    """CX分配くん: インセンティブ計算・管理者専用ページ"""
    # 最高管理者専用
    if st.session_state.get("user_role") != "superadmin":
        st.warning("⛔ この画面は最高管理者のみアクセスできます。")
        return

    # 対象月の決定
    from datetime import datetime
    if target_month:
        month_str = target_month  # YYYYMM
    else:
        month_str = datetime.now().strftime("%Y%m")

    year_label = month_str[:4]
    month_label = str(int(month_str[4:]))

    st.header(f"💰 CX分配くん ── {year_label}年{month_label}月")
    st.caption("インセンティブ分配を計算する管理者専用ツールです。既存の成績ランキングを参照します。")

    # CX設定の読み込み
    cx_config = load_cx_config(month_str)

    # タブ構成
    tab_settings, tab_result = st.tabs(["⚙️ 設定", "📊 計算結果"])

    # ============================================================
    # 設定タブ
    # ============================================================
    with tab_settings:
        st.markdown("### 基本設定")

        col_mode, col_base = st.columns(2)
        with col_mode:
            # 計算モードの切り替え
            mode_options = {
                "A：移行期間モード（旧手当保証＋余剰分配）": "A",
                "B：完全分配モード（全額ポイント分配）": "B",
            }
            current_mode_label = next(
                (k for k, v in mode_options.items() if v == cx_config.get("calculation_mode", "A")),
                "A：移行期間モード（旧手当保証＋余剰分配）"
            )
            selected_mode_label = st.radio(
                "計算モード（フェーズ）",
                list(mode_options.keys()),
                index=list(mode_options.keys()).index(current_mode_label),
                help="モードA: 旧手当を固定保証した上で余剰を分配。モードB: 全額をポイント分配。"
            )
            cx_config["calculation_mode"] = mode_options[selected_mode_label]

        with col_base:
            # 当月CXポイント原資
            cx_config["cx_total_points"] = st.number_input(
                f"当月CXポイント原資（{year_label}年{month_label}月）",
                min_value=0,
                value=int(cx_config.get("cx_total_points", 0)),
                step=1000,
                help="当月の総インセンティブ原資（円）を入力してください。"
            )
            # ベースポイント
            cx_config["base_pt"] = st.number_input(
                "ベースポイント（全員一律）",
                min_value=0,
                value=int(cx_config.get("base_pt", 50)),
                step=1,
                help="全スタッフに付与する基礎ポイント。デフォルトは50pt。"
            )

        st.markdown("---")
        st.markdown("### スキル手当設定（ドコモ支援費：月額）")
        st.caption("各スキル資格手当の月額金額（円）を設定してください。各スタッフに割り当てた区分の金額が旧手当計算に使われます。")

        skill_cols = st.columns(4)
        skill_allowances = cx_config.get("skill_allowances", {})
        skill_names = ["フロントスペシャリスト", "グランマイスター", "マイスター", "プレマイスター"]
        for i, sk in enumerate(skill_names):
            with skill_cols[i]:
                skill_allowances[sk] = st.number_input(
                    sk,
                    min_value=0,
                    value=int(skill_allowances.get(sk, 0)),
                    step=50,
                    key=f"skill_{sk}"
                )
        cx_config["skill_allowances"] = skill_allowances

        st.markdown("---")
        st.markdown("### 成果ポイントテーブル設定")
        st.caption("店長・リーダーの順位別ポイント、スタッフの上下限を設定できます。「1位の値」を変えるとデフォルト比率で他の順位が自動再計算されます。直接入力も可賭。")

        # ---- 店長ポイントテーブル（動的行追加対応） ----
        with st.expander("🏢 店長 成果ポイント（行追加で店舗数増加に対応）", expanded=False):
            st.caption("1位（トップ）を変えるとデフォルト比率で全行が自動計算されます。行の追加・削除も可能です。")

            # session_stateキー（月別にユニーク化）
            _mgr_sk = f"cx_mgr_table_{month_str}"

            # 初回または設定変更時にsession_stateを初期化
            mgr_pts_raw = cx_config.get("manager_pts", {str(k): v for k, v in MANAGER_DEFAULT_PTS.items()})
            if _mgr_sk not in st.session_state:
                st.session_state[_mgr_sk] = pd.DataFrame([
                    {"順位": int(k), "成果pt": int(v)}
                    for k, v in sorted(mgr_pts_raw.items(), key=lambda x: int(x[0]))
                ])

            def _recalc_mgr_dyn():
                """1位トップ値から全行を比率で再計算してsession_stateを更新"""
                top = st.session_state.get("cx_mgr_top_input", 130)
                cur_df = st.session_state[_mgr_sk].copy()
                for i, row in cur_df.iterrows():
                    rk = int(row["順位"])
                    ratio = MANAGER_DEFAULT_RATIOS.get(rk, 0)
                    cur_df.at[i, "成果pt"] = round(top * ratio)
                if not cur_df.empty:
                    cur_df.at[0, "成果pt"] = top  # 1位は必ずtop値
                st.session_state[_mgr_sk] = cur_df

            # 現在の1位の値を取得
            _mgr_cur_top = int(st.session_state[_mgr_sk].iloc[0]["成果pt"]) if not st.session_state[_mgr_sk].empty else 130
            st.number_input(
                "1位（トップ）← ここを変えると比率で全行自動計算",
                min_value=0, value=_mgr_cur_top, step=1,
                key="cx_mgr_top_input",
                on_change=_recalc_mgr_dyn,
            )

            # 動的テーブル（行追加・削除可能）
            edited_mgr_df = st.data_editor(
                st.session_state[_mgr_sk],
                column_config={
                    "順位": st.column_config.NumberColumn("順位", min_value=1, step=1, help="店舗ランキングの順位"),
                    "成果pt": st.column_config.NumberColumn("成果pt", min_value=0, step=1, help="該当順位の成果ポイント"),
                },
                num_rows="dynamic",
                use_container_width=True,
                key=f"cx_mgr_editor_{month_str}",
            )
            # data_editorの編集結果をconfig用dictに変換
            cx_config["manager_pts"] = {
                str(int(row["順位"])): int(row["成果pt"])
                for _, row in edited_mgr_df.dropna(subset=["順位", "成果pt"]).iterrows()
                if pd.notna(row["順位"]) and pd.notna(row["成果pt"])
            }

        # ---- リーダーポイントテーブル（動的行追加対応） ----
        with st.expander("👥 リーダー 成果ポイント（行追加でチーム数増加に対応）", expanded=False):
            st.caption("1位（トップ）を変えるとデフォルト比率で全行が自動計算されます。7位以下はデフォルト0pt。行追加で対応できます。")

            _ldr_sk = f"cx_ldr_table_{month_str}"

            ldr_pts_raw = cx_config.get("leader_pts", {str(k): v for k, v in LEADER_DEFAULT_PTS.items()})
            if _ldr_sk not in st.session_state:
                st.session_state[_ldr_sk] = pd.DataFrame([
                    {"順位": int(k), "成果pt": int(v)}
                    for k, v in sorted(ldr_pts_raw.items(), key=lambda x: int(x[0]))
                ])

            def _recalc_ldr_dyn():
                """1位トップ値から全行を比率で再計算してsession_stateを更新"""
                top = st.session_state.get("cx_ldr_top_input", 120)
                cur_df = st.session_state[_ldr_sk].copy()
                for i, row in cur_df.iterrows():
                    rk = int(row["順位"])
                    ratio = LEADER_DEFAULT_RATIOS.get(rk, 0)
                    cur_df.at[i, "成果pt"] = round(top * ratio)
                if not cur_df.empty:
                    cur_df.at[0, "成果pt"] = top  # 1位は必ずtop値
                st.session_state[_ldr_sk] = cur_df

            _ldr_cur_top = int(st.session_state[_ldr_sk].iloc[0]["成果pt"]) if not st.session_state[_ldr_sk].empty else 120
            st.number_input(
                "1位（トップ）← ここを変えると比率で全行自動計算",
                min_value=0, value=_ldr_cur_top, step=1,
                key="cx_ldr_top_input",
                on_change=_recalc_ldr_dyn,
            )

            edited_ldr_df = st.data_editor(
                st.session_state[_ldr_sk],
                column_config={
                    "順位": st.column_config.NumberColumn("順位", min_value=1, step=1, help="チームランキングの順位"),
                    "成果pt": st.column_config.NumberColumn("成果pt", min_value=0, step=1, help="該当順位の成果ポイント"),
                },
                num_rows="dynamic",
                use_container_width=True,
                key=f"cx_ldr_editor_{month_str}",
            )
            cx_config["leader_pts"] = {
                str(int(row["順位"])): int(row["成果pt"])
                for _, row in edited_ldr_df.dropna(subset=["順位", "成果pt"]).iterrows()
                if pd.notna(row["順位"]) and pd.notna(row["成果pt"])
            }

        # ---- スタッフ 線形補間設定 ----
        with st.expander("👤 スタッフ 成果ポイント（1位上限・最下位下限を設定）", expanded=False):
            st.caption("役職者を除外したスタッフランキングに基づく線形補間の上限（１位）と下限（最下位）を設定します。")
            col_top, col_bottom = st.columns(2)
            with col_top:
                cx_config["member_pt_top"] = st.number_input(
                    "1位（上限 pt）",
                    min_value=0,
                    value=int(cx_config.get("member_pt_top", 100)),
                    step=1,
                    key="cx_member_pt_top",
                    help="スタッフランキング1位の成果ポイント"
                )
            with col_bottom:
                cx_config["member_pt_bottom"] = st.number_input(
                    "最下位（下限 pt）",
                    min_value=0,
                    value=int(cx_config.get("member_pt_bottom", 0)),
                    step=1,
                    key="cx_member_pt_bottom",
                    help="スタッフランキング最下位の成果ポイント"
                )

        st.markdown("---")
        st.markdown("### スタッフ管理")

        # 既存メンバーが空の場合、user_masterから自動転記
        if not cx_config.get("members"):
            cx_config["members"] = get_default_cx_members(user_master)
            st.info("ℹ️ ブライトパスくんのメンバー設定を自動転記しました。役職・スキル手当を確認・修正してください。")

        members = cx_config["members"]

        # スタッフ一覧テーブル（編集可能）
        members_df = pd.DataFrame(members)
        if members_df.empty:
            members_df = pd.DataFrame(columns=["name", "role", "skill_type", "manual_pt"])

        # 表示列の整備
        for col in ["name", "role", "skill_type", "manual_pt"]:
            if col not in members_df.columns:
                members_df[col] = "" if col != "manual_pt" else None

        # 役職を日本語ラベルで表示
        members_df["role_label"] = members_df["role"].map(ROLE_LABELS).fillna("スタッフ")
        display_df = members_df[["name", "role_label", "skill_type", "manual_pt"]].copy()
        display_df.columns = ["スタッフ名", "役職", "スキル手当区分", "手動ポイント(空=自動)"]

        edited_df = st.data_editor(
            display_df,
            column_config={
                "スタッフ名": st.column_config.TextColumn("スタッフ名"),
                "役職": st.column_config.SelectboxColumn(
                    "役職",
                    options=["店長", "リーダー", "スタッフ"],
                    required=True,
                ),
                "スキル手当区分": st.column_config.SelectboxColumn(
                    "スキル手当区分",
                    options=SKILL_TYPES,
                    required=True,
                ),
                "手動ポイント(空=自動)": st.column_config.NumberColumn(
                    "手動ポイント(空=自動)",
                    help="ランキングなしのスタッフに直接ポイントを入力。空欄の場合は自動計算。",
                    min_value=0,
                    step=1,
                ),
            },
            num_rows="dynamic",
            use_container_width=True,
            key="cx_member_editor",
        )

        # 保存ボタン
        if st.button("💾 設定を保存", type="primary", use_container_width=True):
            new_members = []
            for _, row in edited_df.iterrows():
                name = str(row.get("スタッフ名", "")).strip()
                if not name:
                    continue
                role_jp = str(row.get("役職", "スタッフ"))
                role_key = ROLE_LABELS_REV.get(role_jp, "role_member")
                manual_pt_val = row.get("手動ポイント(空=自動)")
                manual_pt = None
                if manual_pt_val is not None and not (isinstance(manual_pt_val, float) and pd.isna(manual_pt_val)):
                    try:
                        manual_pt = int(manual_pt_val)
                    except (ValueError, TypeError):
                        manual_pt = None

                new_members.append({
                    "name": name,
                    "role": role_key,
                    "skill_type": str(row.get("スキル手当区分", "なし")),
                    "manual_pt": manual_pt,
                })
            cx_config["members"] = new_members
            cx_config["month"] = month_str

            if save_cx_config(cx_config, month_str):
                st.success(f"✅ {year_label}年{month_label}月のCX設定を保存しました。")
                st.rerun()
            else:
                st.error("❌ 保存に失敗しました。")

    # ============================================================
    # 計算結果タブ
    # ============================================================
    with tab_result:
        st.markdown(f"### {year_label}年{month_label}月 インセンティブ計算結果")

        # 入力チェック
        if cx_config.get("cx_total_points", 0) <= 0:
            st.warning("⚠️ 設定タブで「当月CXポイント原資」を入力して保存してください。")
            return
        if not cx_config.get("members"):
            st.warning("⚠️ 設定タブでスタッフ情報を保存してください。")
            return

        # 既存のランキングデータを再計算（読み取りのみ・既存ロジックを流用）
        shop_rank_df = None
        team_rank_df = None
        individual_scored_df = None

        if df is not None and not df.empty:
            try:
                config_with_master = dict(config)
                config_with_master["_user_master"] = user_master

                df_sorted = apply_robust_sorting(df, user_master, "スタッフ名", "店舗名", "チーム名")

                df_orders_path = os.path.join(DATA_DIR, "daily_orders.csv")
                df_terms_path = os.path.join(DATA_DIR, "daily_terminals.csv")
                df_orders_cx = pd.read_csv(df_orders_path) if os.path.exists(df_orders_path) else pd.DataFrame()
                df_terms_cx = pd.read_csv(df_terms_path) if os.path.exists(df_terms_path) else pd.DataFrame()
                df_users_cx = pd.DataFrame(user_master.get("users", []))

                df_denom_cx = calculate_denominators(df_sorted, df_orders_cx, df_terms_cx, df_users_cx)
                individual_scored_df = calculate_scores(df_sorted, config_with_master, df_denom_cx)
                team_rank_df = aggregate_team_scores(individual_scored_df, df_sorted, config_with_master)
                shop_rank_df = aggregate_shop_scores(individual_scored_df, df_sorted, config_with_master)
            except Exception as e:
                st.warning(f"⚠️ ランキングデータ取得中にエラーが発生しました: {e}")

        # CX計算実行
        try:
            calc_result = calculate_cx_distribution(
                cx_config,
                shop_rank_df,
                team_rank_df,
                individual_scored_df,
                user_master,
            )
        except Exception as e:
            st.error(f"❌ 計算エラー: {e}")
            return

        mode = calc_result["mode"]
        pt_unit_price = calc_result["pt_unit_price"]
        step1 = calc_result["step1"]
        step3 = calc_result["step3"]
        results = step3["results"]

        # サマリー表示
        mode_label = "モードA（移行期間）" if mode == "A" else "モードB（完全分配）"
        st.markdown(f"**計算モード：** {mode_label}")

        col_s1, col_s2, col_s3, col_s4 = st.columns(4)
        with col_s1:
            st.metric("当月CXポイント原資", f"¥{cx_config['cx_total_points']:,}")
        with col_s2:
            st.metric("固定費総額（旧手当合計）", f"¥{step1['fixed_cost_total']:,}")
        with col_s3:
            st.metric("変動分配用余剰原資", f"¥{int(step1['surplus_fund']):,}")
        with col_s4:
            st.metric("1Ptあたり単価", f"¥{pt_unit_price:,.2f}")

        st.markdown("---")
        st.caption(f"📌 今月の1Pt単価：¥{pt_unit_price:,.2f} / pt　（ベースPt：{cx_config.get('base_pt', 50)}pt）")

        # 詳細一覧テーブル
        rows = []
        for name, r in results.items():
            row = {
                "スタッフ名": name,
                "役職": ROLE_LABELS.get(
                    next((m["role"] for m in cx_config["members"] if m["name"] == name), "role_member"),
                    "スタッフ",
                ),
                "ベースPt": r["base_pt"],
                "成果Pt": r["performance_pt"],
                "合計Pt": r["total_pt"],
                "今月の実力歩合（円）": r["歩合額"],
            }
            if mode == "A":
                row["スキル資格手当（円）"] = r["旧手当"]
                row["最終支給額（円）"] = r["最終支給額"]
            else:
                row["最終支給額（円）"] = r["最終支給額"]

            if r.get("is_manual"):
                row["ベースPt"] = "（手動）"
                row["成果Pt"] = "（手動）"
            rows.append(row)

        result_df = pd.DataFrame(rows)
        st.dataframe(result_df, use_container_width=True, hide_index=True)

        # Excelダウンロード
        try:
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                result_df.to_excel(writer, index=False, sheet_name="CX分配結果")
                summary_data = {
                    "項目": ["当月CXポイント原資", "固定費総額", "変動分配用余剰原資",
                             "システム合計ポイント", "1Pt単価", "計算モード", "ベースポイント"],
                    "値": [
                        f"¥{cx_config['cx_total_points']:,}",
                        f"¥{step1['fixed_cost_total']:,}",
                        f"¥{int(step1['surplus_fund']):,}",
                        f"{step3['system_total_pt']:,} pt",
                        f"¥{pt_unit_price:,.2f}",
                        mode_label,
                        f"{cx_config.get('base_pt', 50)} pt",
                    ],
                }
                pd.DataFrame(summary_data).to_excel(writer, index=False, sheet_name="サマリー")
            buf.seek(0)
            st.download_button(
                label="📥 計算結果をExcelでダウンロード",
                data=buf.getvalue(),
                file_name=f"CX分配結果_{month_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as e:
            st.warning(f"Excelダウンロードの生成に失敗しました: {e}")


if __name__ == "__main__":
    main()
