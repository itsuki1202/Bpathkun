"""
認証管理モジュール
- ユーザー情報（config.yaml）の読み書き
- ログイン履歴（login_log.csv）の記録・読み込み
- パスワードのハッシュ化ユーティリティ
"""

import os
import yaml
import bcrypt
import csv
import pandas as pd
from datetime import datetime

# ファイルパス定義
AUTH_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "auth")
CONFIG_PATH = os.path.join(AUTH_DIR, "config.yaml")
LOG_PATH = os.path.join(AUTH_DIR, "login_log.csv")


def load_auth_config() -> dict:
    """config.yaml を読み込んで返す"""
    if not os.path.exists(CONFIG_PATH):
        # 初期設定が存在しない場合は空の構造を返す
        return {
            "credentials": {"usernames": {}},
            "cookie": {
                "expiry_days": 1,
                "key": "brightpath_secret_key_2026",
                "name": "brightpath_auth"
            },
            "preauthorized": {"emails": []}
        }
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def save_auth_config(config: dict):
    """config.yaml に書き込む"""
    os.makedirs(AUTH_DIR, exist_ok=True)
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        yaml.dump(config, f, allow_unicode=True, default_flow_style=False)


def hash_password(plain_password: str) -> str:
    """パスワードをbcryptでハッシュ化して返す"""
    return bcrypt.hashpw(plain_password.encode(), bcrypt.gensalt()).decode()


def verify_password(plain_password: str, hashed: str) -> bool:
    """パスワードを検証する"""
    return bcrypt.checkpw(plain_password.encode(), hashed.encode())


def get_all_users(config: dict) -> list:
    """ユーザー一覧を取得する [{username, name, email, role}, ...]"""
    users = []
    for uname, udata in config.get("credentials", {}).get("usernames", {}).items():
        users.append({
            "username": uname,
            "name": udata.get("name", ""),
            "email": udata.get("email", ""),
            "role": udata.get("role", "viewer"),
        })
    return users


def add_user(config: dict, username: str, display_name: str, email: str,
             plain_password: str, role: str = "viewer") -> tuple:
    """
    ユーザーを追加する
    Returns: (成功フラグ, メッセージ)
    """
    usernames = config.get("credentials", {}).get("usernames", {})
    if username in usernames:
        return False, f"ユーザーID「{username}」は既に存在します"
    if not username or not plain_password:
        return False, "ユーザーIDとパスワードは必須です"
    usernames[username] = {
        "name": display_name,
        "email": email,
        "password": hash_password(plain_password),
        "role": role,
    }
    config["credentials"]["usernames"] = usernames
    save_auth_config(config)
    return True, f"ユーザー「{username}」を追加しました"


def delete_user(config: dict, username: str) -> tuple:
    """
    ユーザーを削除する
    Returns: (成功フラグ, メッセージ)
    """
    usernames = config.get("credentials", {}).get("usernames", {})
    if username not in usernames:
        return False, f"ユーザーID「{username}」は存在しません"
    if username == "admin":
        return False, "管理者アカウントは削除できません"
    del usernames[username]
    config["credentials"]["usernames"] = usernames
    save_auth_config(config)
    return True, f"ユーザー「{username}」を削除しました"


def reset_password(config: dict, username: str, new_password: str) -> tuple:
    """
    パスワードをリセットする
    Returns: (成功フラグ, メッセージ)
    """
    usernames = config.get("credentials", {}).get("usernames", {})
    if username not in usernames:
        return False, f"ユーザーID「{username}」は存在しません"
    if not new_password:
        return False, "新しいパスワードを入力してください"
    usernames[username]["password"] = hash_password(new_password)
    config["credentials"]["usernames"] = usernames
    save_auth_config(config)
    return True, f"「{username}」のパスワードを変更しました"


def log_login(username: str, display_name: str, status: str, message: str = ""):
    """
    ログイン履歴を login_log.csv に追記する
    status: "success" or "failed"
    """
    os.makedirs(AUTH_DIR, exist_ok=True)
    # ファイルが存在しない場合はヘッダーも書く
    file_exists = os.path.exists(LOG_PATH)
    with open(LOG_PATH, "a", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(["timestamp", "username", "display_name", "status", "message"])
        writer.writerow([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            username,
            display_name,
            status,
            message,
        ])


def load_login_log() -> pd.DataFrame:
    """login_log.csv を DataFrame で返す"""
    if not os.path.exists(LOG_PATH):
        return pd.DataFrame(columns=["timestamp", "username", "display_name", "status", "message"])
    try:
        df = pd.read_csv(LOG_PATH)
        if df.empty:
            return pd.DataFrame(columns=["timestamp", "username", "display_name", "status", "message"])
        df["timestamp"] = pd.to_datetime(df["timestamp"])
        return df.sort_values("timestamp", ascending=False).reset_index(drop=True)
    except Exception:
        return pd.DataFrame(columns=["timestamp", "username", "display_name", "status", "message"])


def generate_user_template_excel() -> bytes:
    """
    ユーザー一括登録用のExcelテンプレートをバイト列で返す。
    列: ユーザーID / 表示名 / 初期パスワード / ロール
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "ユーザー一括登録"

    # ヘッダー定義
    headers = ["ユーザーID", "表示名", "初期パスワード", "ロール"]
    col_widths = [20, 20, 20, 12]
    notes = [
        "英数字のみ。例: tanaka01",
        "画面に表示される名前。例: 田中 太郎",
        "初期パスワード（平文で記入）",
        "viewer または admin"
    ]

    # スタイル定義
    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    note_fill   = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    note_font   = Font(italic=True, color="546E7A", size=9)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center")

    # 1行目: ヘッダー
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_fill and header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    # 2行目: 記入例（注釈）
    for col_idx, note in enumerate(notes, start=1):
        cell = ws.cell(row=2, column=col_idx, value=note)
        cell.fill = note_fill
        cell.font = note_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border

    # 3行目以降: サンプルデータ
    samples = [
        ("tanaka01", "田中 太郎", "initial_pass_01", "viewer"),
        ("yamada02", "山田 花子", "initial_pass_02", "viewer"),
    ]
    for row_idx, sample in enumerate(samples, start=3):
        for col_idx, val in enumerate(sample, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border

    # 行の高さ
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    # バイト列として返す
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def bulk_add_users(config: dict, df: pd.DataFrame) -> tuple:
    """
    DataFrameからユーザーを一括追加する。
    必須列: ユーザーID / 表示名 / 初期パスワード / ロール
    Returns: (成功数, スキップ数, メッセージリスト)
    """
    usernames = config.get("credentials", {}).get("usernames", {})
    success_count  = 0
    skip_count     = 0
    messages       = []
    valid_roles    = {"viewer", "admin"}

    for _, row in df.iterrows():
        uid  = str(row.get("ユーザーID", "")).strip()
        name = str(row.get("表示名", "")).strip()
        pw   = str(row.get("初期パスワード", "")).strip()
        role = str(row.get("ロール", "viewer")).strip().lower()

        # 空行スキップ
        if not uid or uid == "nan" or not pw or pw == "nan":
            skip_count += 1
            messages.append(f"  ⚠️ スキップ（空行）: UID='{uid}'")
            continue

        # 注釈行スキップ（英数字のみ ... など）
        if uid in ("ユーザーID", "英数字のみ。例: tanaka01"):
            skip_count += 1
            continue

        # 重複スキップ
        if uid in usernames:
            skip_count += 1
            messages.append(f"  ⚠️ スキップ（重複）: UID='{uid}'")
            continue

        # ロール正規化
        if role not in valid_roles:
            role = "viewer"

        usernames[uid] = {
            "name":     name if name and name != "nan" else uid,
            "email":    "",
            "password": hash_password(pw),
            "role":     role,
        }
        success_count += 1
        messages.append(f"  ✅ 登録: UID='{uid}' / 名前='{name}' / ロール={role}")

    config["credentials"]["usernames"] = usernames
    save_auth_config(config)
    return success_count, skip_count, messages

